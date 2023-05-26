#-- codingutf-8 --

def ErrorLog(error: str):
    current_time = time.strftime("%Y.%m.%d/%H:%M:%S", time.localtime(time.time()))
    path = "\\\\192.168.120.85\\vps공유\\00. VP승인 프로그램모음\\13.개발프로그램\\로그\\Log.txt"
    with open(path, "a") as f:
        f.write(f"[{current_time}] - {error}\n")

try:

    from PyQt5.QtWidgets import *
    from PyQt5 import uic
    from PyQt5.QtGui import *
    from PyQt5.QtCore import *
    import sys,pymysql,time,traceback
    from openpyxl import load_workbook
    import pyautogui

    sys.path.append('resource\\py\\')

    import setting

    ip = setting.setting.ip()
    password = setting.setting.password()

    form_secondclass = uic.loadUiType("resource/ui/calendar.ui")[0]   # UI 파일 불러와서 변수에 저장

    class MyWindow(QMainWindow):   # 메인윈도우 클래스 시작

        def addtabwidget(widget,tab,onenum):   # addtabwidget 함수 시작

            try:

                global pushButton,pushButton_7,tableWidget,lineEdit_4,lineEdit,lineEdit_2
                global lineEdit_7,lineEdit_8,textEdit_2,lineEdit_5,lineEdit_6,lineEdit_3
                global textEdit,pushButton_12,lineEdit_9,checkBox,glotab,glowidget,gloonenum

                glowidget = widget
                glotab = tab
                gloonenum = onenum

                centralwidget = QWidget()
                centralwidget.setObjectName("centralwidget")
                gridLayout_3 = QGridLayout(centralwidget)
                gridLayout_3.setContentsMargins(0, 0, 0, 0)
                gridLayout_3.setSpacing(6)
                gridLayout_3.setObjectName("gridLayout_3")
                gridLayout = QGridLayout()
                gridLayout.setObjectName("gridLayout")
                pushButton_12 = QPushButton(centralwidget)
                pushButton_12.setMinimumSize(QSize(0, 30))
                pushButton_12.setObjectName("pushButton_12")
                gridLayout.addWidget(pushButton_12, 2, 3, 1, 1)
                textEdit = QTextEdit(centralwidget)
                textEdit.setMinimumSize(QSize(0, 60))
                font = QFont()
                font.setFamily("맑은 고딕")
                textEdit.setFont(font)
                textEdit.setStyleSheet("background: #DEEBF7;")
                textEdit.setObjectName("textEdit")
                gridLayout.addWidget(textEdit, 1, 1, 1, 3)
                label = QLabel(centralwidget)
                sizePolicy = QSizePolicy(QSizePolicy.Preferred, QSizePolicy.Preferred)
                sizePolicy.setHorizontalStretch(0)
                sizePolicy.setVerticalStretch(0)
                sizePolicy.setHeightForWidth(label.sizePolicy().hasHeightForWidth())
                label.setSizePolicy(sizePolicy)
                label.setMinimumSize(QSize(60, 20))
                label.setAlignment(Qt.AlignCenter)
                label.setObjectName("label")
                gridLayout.addWidget(label, 1, 0, 1, 1)
                spacerItem = QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)
                gridLayout.addItem(spacerItem, 2, 1, 1, 1)
                label_7 = QLabel(centralwidget)
                sizePolicy = QSizePolicy(QSizePolicy.Preferred, QSizePolicy.Preferred)
                sizePolicy.setHorizontalStretch(0)
                sizePolicy.setVerticalStretch(0)
                sizePolicy.setHeightForWidth(label_7.sizePolicy().hasHeightForWidth())
                label_7.setSizePolicy(sizePolicy)
                label_7.setMinimumSize(QSize(60, 20))
                label_7.setAlignment(Qt.AlignCenter)
                label_7.setObjectName("label_7")
                gridLayout.addWidget(label_7, 0, 0, 1, 1)
                pushButton_7 = QPushButton(centralwidget)
                pushButton_7.setMinimumSize(QSize(0, 30))
                pushButton_7.setObjectName("pushButton_7")
                gridLayout.addWidget(pushButton_7, 2, 2, 1, 1)
                textEdit_2 = QTextEdit(centralwidget)
                textEdit_2.setMinimumSize(QSize(0, 60))
                textEdit_2.setReadOnly(True)
                textEdit_2.setObjectName("textEdit_2")
                gridLayout.addWidget(textEdit_2, 0, 1, 1, 3)
                gridLayout_3.addLayout(gridLayout, 1, 0, 1, 1)
                tableWidget = QTableWidget(centralwidget)
                tableWidget.setObjectName("tableWidget")
                tableWidget.setColumnCount(0)
                tableWidget.setRowCount(0)
                gridLayout_3.addWidget(tableWidget, 2, 0, 1, 1)
                gridLayout_2 = QGridLayout()
                gridLayout_2.setObjectName("gridLayout_2")
                verticalLayout_8 = QVBoxLayout()
                verticalLayout_8.setObjectName("verticalLayout_8")
                lineEdit_6 = QLineEdit(centralwidget)
                font = QFont()
                font.setFamily("맑은 고딕")
                lineEdit_6.setFont(font)
                lineEdit_6.setStyleSheet("background: #DEEBF7;")
                lineEdit_6.setObjectName("lineEdit_6")
                verticalLayout_8.addWidget(lineEdit_6)
                lineEdit_7 = QLineEdit(centralwidget)
                lineEdit_7.setReadOnly(True)
                lineEdit_7.setObjectName("lineEdit_7")
                verticalLayout_8.addWidget(lineEdit_7)
                lineEdit_8 = QLineEdit(centralwidget)
                lineEdit_8.setReadOnly(True)
                lineEdit_8.setObjectName("lineEdit_8")
                verticalLayout_8.addWidget(lineEdit_8)
                gridLayout_2.addLayout(verticalLayout_8, 2, 6, 1, 1)
                verticalLayout_2 = QVBoxLayout()
                verticalLayout_2.setObjectName("verticalLayout_2")
                label_4 = QLabel(centralwidget)
                label_4.setMinimumSize(QSize(60, 20))
                label_4.setAlignment(Qt.AlignCenter)
                label_4.setObjectName("label_4")
                verticalLayout_2.addWidget(label_4)
                label_5 = QLabel(centralwidget)
                label_5.setMinimumSize(QSize(60, 20))
                label_5.setAlignment(Qt.AlignCenter)
                label_5.setObjectName("label_5")
                verticalLayout_2.addWidget(label_5)
                label_17 = QLabel(centralwidget)
                label_17.setMinimumSize(QSize(60, 20))
                font = QFont()
                font.setStrikeOut(False)
                label_17.setFont(font)
                label_17.setAlignment(Qt.AlignCenter)
                label_17.setObjectName("label_17")
                verticalLayout_2.addWidget(label_17)
                gridLayout_2.addLayout(verticalLayout_2, 2, 2, 1, 1)
                verticalLayout_13 = QVBoxLayout()
                verticalLayout_13.setObjectName("verticalLayout_13")
                pushButton_9 = QPushButton(centralwidget)
                sizePolicy = QSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)
                sizePolicy.setHorizontalStretch(0)
                sizePolicy.setVerticalStretch(0)
                sizePolicy.setHeightForWidth(pushButton_9.sizePolicy().hasHeightForWidth())
                pushButton_9.setSizePolicy(sizePolicy)
                pushButton_9.setMinimumSize(QSize(20, 20))
                pushButton_9.setMaximumSize(QSize(20, 20))
                pushButton_9.setObjectName("pushButton_9")
                verticalLayout_13.addWidget(pushButton_9)
                pushButton_10 = QPushButton(centralwidget)
                sizePolicy = QSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)
                sizePolicy.setHorizontalStretch(0)
                sizePolicy.setVerticalStretch(0)
                sizePolicy.setHeightForWidth(pushButton_10.sizePolicy().hasHeightForWidth())
                pushButton_10.setSizePolicy(sizePolicy)
                pushButton_10.setMinimumSize(QSize(20, 20))
                pushButton_10.setMaximumSize(QSize(20, 20))
                pushButton_10.setObjectName("pushButton_10")
                verticalLayout_13.addWidget(pushButton_10)
                pushButton_11 = QPushButton(centralwidget)
                sizePolicy = QSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)
                sizePolicy.setHorizontalStretch(0)
                sizePolicy.setVerticalStretch(0)
                sizePolicy.setHeightForWidth(pushButton_11.sizePolicy().hasHeightForWidth())
                pushButton_11.setSizePolicy(sizePolicy)
                pushButton_11.setMinimumSize(QSize(20, 20))
                pushButton_11.setMaximumSize(QSize(20, 20))
                pushButton_11.setObjectName("pushButton_11")
                verticalLayout_13.addWidget(pushButton_11)
                gridLayout_2.addLayout(verticalLayout_13, 2, 7, 1, 1)
                verticalLayout_3 = QVBoxLayout()
                verticalLayout_3.setObjectName("verticalLayout_3")
                lineEdit = QLineEdit(centralwidget)
                lineEdit.setReadOnly(True)
                lineEdit.setObjectName("lineEdit")
                verticalLayout_3.addWidget(lineEdit)
                lineEdit_2 = QLineEdit(centralwidget)
                lineEdit_2.setReadOnly(True)
                lineEdit_2.setObjectName("lineEdit_2")
                verticalLayout_3.addWidget(lineEdit_2)
                checkBox = QCheckBox(centralwidget)
                checkBox.setObjectName("checkBox")
                verticalLayout_3.addWidget(checkBox)
                gridLayout_2.addLayout(verticalLayout_3, 2, 1, 1, 1)
                verticalLayout_11 = QVBoxLayout()
                verticalLayout_11.setObjectName("verticalLayout_11")
                pushButton_3 = QPushButton(centralwidget)
                sizePolicy = QSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)
                sizePolicy.setHorizontalStretch(0)
                sizePolicy.setVerticalStretch(0)
                sizePolicy.setHeightForWidth(pushButton_3.sizePolicy().hasHeightForWidth())
                pushButton_3.setSizePolicy(sizePolicy)
                pushButton_3.setMinimumSize(QSize(20, 20))
                pushButton_3.setMaximumSize(QSize(20, 20))
                pushButton_3.setObjectName("pushButton_3")
                verticalLayout_11.addWidget(pushButton_3)
                pushButton_8 = QPushButton(centralwidget)
                sizePolicy = QSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)
                sizePolicy.setHorizontalStretch(0)
                sizePolicy.setVerticalStretch(0)
                sizePolicy.setHeightForWidth(pushButton_8.sizePolicy().hasHeightForWidth())
                pushButton_8.setSizePolicy(sizePolicy)
                pushButton_8.setMinimumSize(QSize(20, 20))
                pushButton_8.setMaximumSize(QSize(20, 20))
                pushButton_8.setObjectName("pushButton_8")
                verticalLayout_11.addWidget(pushButton_8)
                pushButton_13 = QPushButton(centralwidget)
                sizePolicy = QSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)
                sizePolicy.setHorizontalStretch(0)
                sizePolicy.setVerticalStretch(0)
                sizePolicy.setHeightForWidth(pushButton_13.sizePolicy().hasHeightForWidth())
                pushButton_13.setSizePolicy(sizePolicy)
                pushButton_13.setMinimumSize(QSize(20, 20))
                pushButton_13.setMaximumSize(QSize(20, 20))
                pushButton_13.setObjectName("pushButton_13")
                verticalLayout_11.addWidget(pushButton_13)
                gridLayout_2.addLayout(verticalLayout_11, 2, 4, 1, 1)
                verticalLayout = QVBoxLayout()
                verticalLayout.setObjectName("verticalLayout")
                label_2 = QLabel(centralwidget)
                label_2.setMinimumSize(QSize(60, 20))
                label_2.setAlignment(Qt.AlignCenter)
                label_2.setObjectName("label_2")
                verticalLayout.addWidget(label_2)
                label_3 = QLabel(centralwidget)
                label_3.setMinimumSize(QSize(60, 20))
                label_3.setAlignment(Qt.AlignCenter)
                label_3.setObjectName("label_3")
                verticalLayout.addWidget(label_3)
                label_18 = QLabel(centralwidget)
                label_18.setMinimumSize(QSize(0, 20))
                label_18.setText("")
                label_18.setObjectName("label_18")
                verticalLayout.addWidget(label_18)
                gridLayout_2.addLayout(verticalLayout, 2, 0, 1, 1)
                verticalLayout_4 = QVBoxLayout()
                verticalLayout_4.setObjectName("verticalLayout_4")
                lineEdit_4 = QLineEdit(centralwidget)
                font = QFont()
                font.setFamily("맑은 고딕")
                lineEdit_4.setFont(font)
                lineEdit_4.setReadOnly(True)
                lineEdit_4.setObjectName("lineEdit_4")
                verticalLayout_4.addWidget(lineEdit_4)
                lineEdit_5 = QLineEdit(centralwidget)
                lineEdit_5.setReadOnly(True)
                lineEdit_5.setObjectName("lineEdit_5")
                verticalLayout_4.addWidget(lineEdit_5)
                lineEdit_9 = QLineEdit(centralwidget)
                font = QFont()
                font.setFamily("맑은 고딕")
                lineEdit_9.setFont(font)
                lineEdit_9.setStyleSheet("background: #DEEBF7;")
                lineEdit_9.setObjectName("lineEdit_9")
                verticalLayout_4.addWidget(lineEdit_9)
                gridLayout_2.addLayout(verticalLayout_4, 2, 3, 1, 1)
                verticalLayout_7 = QVBoxLayout()
                verticalLayout_7.setObjectName("verticalLayout_7")
                label_6 = QLabel(centralwidget)
                label_6.setMinimumSize(QSize(60, 20))
                label_6.setAlignment(Qt.AlignCenter)
                label_6.setObjectName("label_6")
                verticalLayout_7.addWidget(label_6)
                label_15 = QLabel(centralwidget)
                label_15.setMinimumSize(QSize(60, 20))
                label_15.setAlignment(Qt.AlignCenter)
                label_15.setObjectName("label_15")
                verticalLayout_7.addWidget(label_15)
                label_16 = QLabel(centralwidget)
                label_16.setMinimumSize(QSize(60, 20))
                label_16.setAlignment(Qt.AlignCenter)
                label_16.setObjectName("label_16")
                verticalLayout_7.addWidget(label_16)
                gridLayout_2.addLayout(verticalLayout_7, 2, 5, 1, 1)
                horizontalLayout_2 = QHBoxLayout()
                horizontalLayout_2.setObjectName("horizontalLayout_2")
                label_8 = QLabel(centralwidget)
                label_8.setMinimumSize(QSize(60, 20))
                label_8.setAlignment(Qt.AlignCenter)
                label_8.setObjectName("label_8")
                horizontalLayout_2.addWidget(label_8)
                lineEdit_3 = QLineEdit(centralwidget)
                lineEdit_3.setReadOnly(True)
                lineEdit_3.setObjectName("lineEdit_3")
                horizontalLayout_2.addWidget(lineEdit_3)
                pushButton_4 = QPushButton(centralwidget)
                pushButton_4.setMinimumSize(QSize(0, 30))
                pushButton_4.setObjectName("pushButton_4")
                horizontalLayout_2.addWidget(pushButton_4)
                gridLayout_2.addLayout(horizontalLayout_2, 1, 0, 1, 8)
                horizontalLayout = QHBoxLayout()
                horizontalLayout.setObjectName("horizontalLayout")
                spacerItem1 = QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)
                horizontalLayout.addItem(spacerItem1)
                pushButton_2 = QPushButton(centralwidget)
                pushButton_2.setMinimumSize(QSize(0, 30))
                pushButton_2.setObjectName("pushButton_2")
                horizontalLayout.addWidget(pushButton_2)
                pushButton = QPushButton(centralwidget)
                pushButton.setMinimumSize(QSize(0, 30))
                pushButton.setObjectName("pushButton")
                horizontalLayout.addWidget(pushButton)
                gridLayout_2.addLayout(horizontalLayout, 0, 0, 1, 8)
                gridLayout_3.addLayout(gridLayout_2, 0, 0, 1, 1)
                gridLayout_3.setRowStretch(0, 1)
                gridLayout_3.setRowStretch(1, 1)
                gridLayout_3.setRowStretch(2, 3)

                centralwidget.setTabOrder(lineEdit_9, lineEdit_6)
                centralwidget.setTabOrder(lineEdit_6, textEdit)
                centralwidget.setTabOrder(textEdit, pushButton_2)
                centralwidget.setTabOrder(pushButton_2, pushButton)
                centralwidget.setTabOrder(pushButton, pushButton_4)
                centralwidget.setTabOrder(pushButton_4, pushButton_3)
                centralwidget.setTabOrder(pushButton_3, pushButton_8)
                centralwidget.setTabOrder(pushButton_8, pushButton_13)
                centralwidget.setTabOrder(pushButton_13, pushButton_9)
                centralwidget.setTabOrder(pushButton_9, pushButton_10)
                centralwidget.setTabOrder(pushButton_10, pushButton_11)
                centralwidget.setTabOrder(pushButton_11, pushButton_7)
                centralwidget.setTabOrder(pushButton_7, pushButton_12)
                centralwidget.setTabOrder(pushButton_12, lineEdit_3)
                centralwidget.setTabOrder(lineEdit_3, lineEdit)
                centralwidget.setTabOrder(lineEdit, lineEdit_2)
                centralwidget.setTabOrder(lineEdit_2, checkBox)
                centralwidget.setTabOrder(checkBox, textEdit_2)
                centralwidget.setTabOrder(textEdit_2, lineEdit_4)
                centralwidget.setTabOrder(lineEdit_4, lineEdit_5)
                centralwidget.setTabOrder(lineEdit_5, lineEdit_7)
                centralwidget.setTabOrder(lineEdit_7, lineEdit_8)
                centralwidget.setTabOrder(lineEdit_8, tableWidget)

                _translate = QCoreApplication.translate
                pushButton_12.setText(_translate("MainWindow", "리스트 초기화"))
                label.setText(_translate("MainWindow", "비    고"))
                label_7.setText(_translate("MainWindow", "작  업  명"))
                pushButton_7.setText(_translate("MainWindow", "리스트 불러오기"))
                label_4.setText(_translate("MainWindow", "작업 지시일"))
                label_5.setText(_translate("MainWindow", "전개 완료일"))
                label_17.setText(_translate("MainWindow", "프레임 완료일"))
                pushButton_9.setText(_translate("MainWindow", "▼"))
                pushButton_10.setText(_translate("MainWindow", "▼"))
                pushButton_11.setText(_translate("MainWindow", "▼"))
                checkBox.setText(_translate("MainWindow", "작업 취소"))
                pushButton_3.setText(_translate("MainWindow", "▼"))
                pushButton_8.setText(_translate("MainWindow", "▼"))
                pushButton_13.setText(_translate("MainWindow", "▼"))
                label_2.setText(_translate("MainWindow", "관리 번호"))
                label_3.setText(_translate("MainWindow", "업  체  명"))
                label_6.setText(_translate("MainWindow", "판금 완료일"))
                label_15.setText(_translate("MainWindow", "생산 완료일"))
                label_16.setText(_translate("MainWindow", "고객 납기일"))
                label_8.setText(_translate("MainWindow", "PDF 경로"))
                pushButton_4.setText(_translate("MainWindow", "경로설정"))
                pushButton_2.setText(_translate("MainWindow", "저장"))
                pushButton.setText(_translate("MainWindow", "닫기"))

                now1 = time.localtime()
                s1 = "%04d%02d%02d" % (now1.tm_year, now1.tm_mon, now1.tm_mday)

                lineEdit_4.setInputMask('0000-00-00')
                lineEdit_4.setText(s1)
                lineEdit_5.setInputMask('0000-00-00')
                lineEdit_5.setText(s1)
                lineEdit_6.setInputMask('0000-00-00')
                lineEdit_6.setText(s1)
                lineEdit_7.setInputMask('0000-00-00')
                lineEdit_7.setText(s1)
                lineEdit_8.setInputMask('0000-00-00')
                lineEdit_8.setText(s1)
                lineEdit_9.setInputMask('0000-00-00')
                lineEdit_9.setText(s1)

                lineEdit.setTextMargins(5,0,0,0)
                lineEdit_2.setTextMargins(5,0,0,0)
                lineEdit_3.setTextMargins(5,0,0,0)
                lineEdit_4.setTextMargins(5,0,0,0)
                lineEdit_5.setTextMargins(5,0,0,0)
                lineEdit_6.setTextMargins(5,0,0,0)
                lineEdit_7.setTextMargins(5,0,0,0)
                lineEdit_8.setTextMargins(5,0,0,0)
                lineEdit_9.setTextMargins(5,0,0,0)
                textEdit.setContentsMargins(5,0,0,0)
                textEdit_2.setContentsMargins(5,0,0,0)

                widget.setLayout(gridLayout_3)

                pushButton.clicked.connect(MyWindow.closecheck)
                pushButton_7.clicked.connect(MyWindow.push7)
                pushButton_3.clicked.connect(MyWindow.dateclick1)
                pushButton_8.clicked.connect(MyWindow.dateclick2)
                pushButton_9.clicked.connect(MyWindow.dateclick3)
                pushButton_10.clicked.connect(MyWindow.dateclick4)
                pushButton_11.clicked.connect(MyWindow.dateclick5)
                pushButton_13.clicked.connect(MyWindow.dateclick6)
                pushButton_4.clicked.connect(MyWindow.push4)
                pushButton_2.clicked.connect(MyWindow.savecheck)
                pushButton_12.clicked.connect(MyWindow.clearcheck)

                global dic

                dic = {}

            except Exception:
                QMessageBox.critical(glowidget,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def edittable():

            try:

                conn = pymysql.connect(host=ip, user='user', password=password, db='vps_planner', charset='utf8', port=1980)   # 데이터 베이스 접속 내용을 conn 변수에 저장

                sql = "SELECT 통합진행번호 FROM planner WHERE 관리번호 = %s"   # 데이터베이스 명령어를 spl 변수에 저장
                with conn.cursor() as cur:
                    cur.execute(sql,gloonenum)
                    allnum = cur.fetchone()
                
                sql = "SELECT * FROM sidedater WHERE 통합진행번호 = %s"   # 데이터베이스 명령어를 spl 변수에 저장
                with conn.cursor() as cur:
                    cur.execute(sql,allnum)
                    data = cur.fetchall()[0]

                lineEdit_3.setText(data[1])
                lineEdit.setText(data[2])
                lineEdit_2.setText(data[3])
                lineEdit_4.setText(str(data[4]))
                lineEdit_5.setText(str(data[5]))
                lineEdit_6.setText(str(data[7]))
                lineEdit_7.setText(str(data[8]))
                lineEdit_8.setText(str(data[9]))
                lineEdit_9.setText(str(data[6]))
                textEdit_2.setText(data[10])
                textEdit.setText(data[11])

                tableWidget.setColumnCount(len(data[18].split('@')))   # tableWidget 테이블 위젯 열 갯수를 headerlist 리스트 갯수만큼 설정
                tableWidget.setHorizontalHeaderLabels(data[18].split('@'))   # tableWidget 테이블 위젯 열 헤더에 headerlist 리스트 입력
                tableWidget.setAutoScroll(True)   # tableWidget 테이블 위젯 내용이 많아질시 자동 스크롤 생성 허용
                tableWidget.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)   # tableWidget 테이블 위젯 수직 스크롤단위를 픽셀단위로 변경
                tableWidget.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)   # tableWidget 테이블 위젯 수평 스크롤단위를 픽셀단위로 변경
                tableWidget.setEditTriggers(QAbstractItemView.EditTriggers(False))   # tableWidget 테이블 위젯 셀 수정 불가하게 설정
                tableWidget.setSelectionBehavior(QAbstractItemView.SelectRows)   # tableWidget 테이블 위젯 셀 선택시 행 전체 선택 설정
                tableWidget.setAlternatingRowColors(True)   # tableWidget 테이블 위젯 연속되는 행 색상 다르게 설정
                tableWidget.setRowCount(int(len(data[19].split('@'))/len(data[18].split('@'))))

                num = 0

                for i in range(int(len(data[19].split('@'))/len(data[18].split('@')))):
                    for j in range(len(data[18].split('@'))):
                        tableWidget.setItem(i,j,QTableWidgetItem(data[19].split('@')[num]))
                        tableWidget.item(i,j).setTextAlignment(Qt.AlignCenter)
                        num += 1

                tableWidget.verticalHeader().hide()
                tableWidget.resizeColumnsToContents()
                tableWidget.resizeRowsToContents()

                if lineEdit != None:

                    sql = "SELECT 취소 FROM sidedater WHERE 관리번호=%s" 
                    with conn.cursor() as cur:
                        cur.execute(sql, (lineEdit.text()))
                        cancle = cur.fetchone()[0]

                    if cancle == '취소':
                        checkBox.setChecked(True)
                    else:
                        checkBox.setChecked(False)

                conn.close()

            except Exception:
                QMessageBox.critical(glowidget,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def push12():

            try:

                now1 = time.localtime()
                s1 = "%04d%02d%02d" % (now1.tm_year, now1.tm_mon, now1.tm_mday)

                lineEdit.clear()
                lineEdit_2.clear()
                textEdit_2.clear()
                textEdit.clear()
                tableWidget.clear()
                tableWidget.setColumnCount(0)
                tableWidget.setRowCount(0)
                lineEdit_4.setText(s1)
                lineEdit_5.setText(s1)
                lineEdit_6.setText(s1)
                lineEdit_7.setText(s1)
                lineEdit_8.setText(s1)

                global dic

                dic = {}

            except Exception:
                QMessageBox.critical(glowidget,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def clearcheck():

            try:
        
                if tableWidget.item(0,0) != None:
                    msgbox = QMessageBox.question(glowidget,'확  인','리스트를 초기화 하시겠습니까?')
                    if msgbox == QMessageBox.Yes:
                        MyWindow.push12()
                        QMessageBox.information(glowidget,'초기화','리스트가 초기화 되었습니다.')
                    elif msgbox == QMessageBox.No:
                        pass
                else:
                    pass

            except Exception:
                QMessageBox.critical(glowidget,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def push2():

            try:

                if checkBox.checkState() == 2:
                    
                    cancle = '취소'
                else:
                    cancle = ''

                conn = pymysql.connect(host=ip, user='user', password=password, db='vps_planner', charset='utf8', port=1980)   # 데이터 베이스 접속 내용을 conn 변수에 저장
                sql = "SELECT MAX(넘버) FROM planner"   # 데이터베이스 명령어를 spl 변수에 저장
                with conn.cursor() as cur:
                    cur.execute(sql)
                num = int(cur.fetchall()[0][0])

                tablehead = []
                for i in range(tableWidget.columnCount()):
                    tablehead.append(tableWidget.horizontalHeaderItem(i).text())
                
                table = []
                for x in range(tableWidget.rowCount()):
                    for y in range(tableWidget.columnCount()):
                        table.append(tableWidget.item(x,y).text())

                pdf = lineEdit_3.text()
                allnum = lineEdit.text()
                allcom = lineEdit_2.text()
                allcre = lineEdit_4.text()
                allufd = lineEdit_5.text()
                allped = lineEdit_6.text()
                allend = lineEdit_7.text()
                alldeli = lineEdit_8.text()
                frend = lineEdit_9.text()
                allwork = textEdit_2.toPlainText()
                alloth = textEdit.toPlainText()
                allhead = '@'.join(tablehead)
                alltable = '@'.join(table)

                if dic != {}:

                    keylist = list(sorted(dic.keys(),reverse=False))
                    ufd = lineEdit_5.text()
                    ped = lineEdit_6.text()
                    oth = textEdit.toPlainText()

                    for i in range(len(keylist)):

                        num += 1
                        key = keylist[i]
                        value = dic.get(key)

                        sql = "SELECT EXISTS(SELECT * FROM planner WHERE 관리번호 = %s)"   # 데이터베이스 명령어를 spl 변수에 저장
                        with conn.cursor() as cur:
                            cur.execute(sql,keylist[i])
                            same = cur.fetchone()[0]

                        if same:

                            msgbox = QMessageBox.question(glowidget,'중복 확인','중복 데이터가 있습니다.\n덮어씌우시겠습니까?\n관리번호 : ' + key)

                            if msgbox == QMessageBox.Yes:

                                for j in range(len(value)):

                                    if j == 0:
                                        cre = str(value[j]).split(' ')[0]
                                    elif j == 1:
                                        com = value[j]
                                    elif j == 4:
                                        work = value[j]
                                    else:
                                        pass

                                if cancle == '취소':
                                    sql = "UPDATE planner SET 업체명=%s,수주요약=%s,작업지시일=%s,전개완료일=%s,프레임완료일=%s,판금완료일=%s,비고=%s,작업현황=%s,통합진행번호=%s WHERE 관리번호=%s"
                                    with conn.cursor() as cur:
                                        cur.execute(sql, (com,work,cre,ufd,frend,ped,oth,'작업취소',keylist[0],key))
                                        conn.commit()
                                else:
                                    sql = "UPDATE planner SET 업체명=%s,수주요약=%s,작업지시일=%s,전개완료일=%s,프레임완료일=%s,판금완료일=%s,비고=%s,작업현황=%s,통합진행번호=%s WHERE 관리번호=%s"
                                    with conn.cursor() as cur:
                                        cur.execute(sql, (com,work,cre,ufd,frend,ped,oth,'작업대기',keylist[0],key))
                                        conn.commit()

                                num -= 1

                            sql = "UPDATE sidedater SET PDF경로=%s,관리번호=%s,업체명=%s,작업지시일=%s,전개완료일=%s,프레임완료일=%s,판금완료일=%s,생산완료일=%s,납품일=%s,작업명=%s,비고=%s,테이블열헤더=%s,테이블내용=%s,취소=%s WHERE 통합진행번호=%s" 
                            with conn.cursor() as cur:
                                cur.execute(sql, (pdf,allnum,allcom,allcre,allufd,frend,allped,allend,alldeli,allwork,alloth,allhead,alltable,cancle,keylist[0]))
                                conn.commit()

                            if msgbox == QMessageBox.No:

                                num -= 1

                        else:

                            for j in range(len(value)):

                                if j == 0:
                                    cre = str(value[j]).split(' ')[0]
                                elif j == 1:
                                    com = value[j]
                                elif j == 4:
                                    work = value[j]
                                else:
                                    pass

                            sql = "INSERT INTO planner VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                            with conn.cursor() as cur:
                                cur.execute(sql, (num,key,com,work,cre,ufd,frend,ped,oth,'0000-00-00 00:00:00','작업대기',keylist[0]))
                                conn.commit()

                            sql = "INSERT INTO sidedater VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                            with conn.cursor() as cur:
                                cur.execute(sql, (keylist[0],pdf,allnum,allcom,allcre,allufd,frend,allped,allend,alldeli,allwork,alloth,'0/0','0/0','0/0','0/0','0/0','00:00:00',allhead,alltable,cancle))
                                conn.commit()

                    conn.close()
                    if keylist != 1:
                        glotab.setTabText(glotab.currentIndex(),keylist[0]+' ...')
                    else:
                        glotab.setTabText(glotab.currentIndex(),keylist[0])

                elif dic == {} and tableWidget.item(0,0) != None:

                    pdf = lineEdit_3.text()
                    ufd = lineEdit_5.text()
                    ped = lineEdit_6.text()
                    oth = textEdit.toPlainText()

                    conn = pymysql.connect(host=ip, user='user', password=password, db='vps_planner', charset='utf8', port=1980)   # 데이터 베이스 접속 내용을 conn 변수에 저장

                    for x in lineEdit.text().split(','):

                        
                        if cancle == '취소':
                            sql = "UPDATE planner SET 전개완료일=%s,판금완료일=%s,비고=%s,작업현황=%s WHERE 관리번호=%s"   # 데이터베이스 명령어를 spl 변수에 저장
                            with conn.cursor() as cur:
                                cur.execute(sql,(ufd,ped,oth,'작업취소',x))
                                conn.commit()

                        else:

                            sql = "UPDATE planner SET 전개완료일=%s,판금완료일=%s,비고=%s,작업현황=%s WHERE 관리번호=%s"   # 데이터베이스 명령어를 spl 변수에 저장
                            with conn.cursor() as cur:
                                cur.execute(sql,(ufd,ped,oth,x,'작업대기'))
                                conn.commit()
                        
                        sql = "UPDATE sidedater SET PDF경로=%s,전개완료일=%s,판금완료일=%s,비고=%s,테이블열헤더=%s,테이블내용=%s,취소=%s WHERE 관리번호=%s"   # 데이터베이스 명령어를 spl 변수에 저장
                        with conn.cursor() as cur:
                            cur.execute(sql,(pdf,ufd,ped,oth,allhead,alltable,cancle,lineEdit.text()))
                            conn.commit()

                    conn.close()
                
                else:
                    pass

            except Exception:
                QMessageBox.critical(glowidget,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))
            
        def savecheck():

            try:
        
                if lineEdit.text() != '':
                    msgbox = QMessageBox.question(glowidget,'저  장','저장하시겠습니까?')
                    if msgbox == QMessageBox.Yes:
                        if lineEdit_3.text() == '':
                            QMessageBox.warning(glowidget,'경  고','PDF 경로가 지정되지 않았습니다.')
                        else:
                            MyWindow.push2()
                            QMessageBox.information(glowidget,'저  장','저장되었습니다.')
                            glotab.removeTab(glotab.currentIndex())
                    elif msgbox == QMessageBox.No:
                        pass
                else:
                    pass

            except Exception:
                QMessageBox.critical(glowidget,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def push4():

            try:

                filesave = QFileDialog.getExistingDirectory(glowidget, 'PDF 경로')
                if '000통합프로그램000' in filesave:
                    filesavespl = filesave.split('/000통합프로그램000/')
                    savepath = '\\\\192.168.120.85\\vpdata\\PROJECT\\000통합프로그램000\\'+filesavespl[1].replace('/','\\')+'\\'
                elif '000통합프로그램000' not in filesave:
                    if 'PROJECT' in filesave:
                        filesavespl = filesave.split('/PROJECT/')
                        savepath = '\\\\192.168.120.85\\vpdata\\PROJECT\\'+filesavespl[1].replace('/','\\')+'\\'
                    elif filesave == '':
                        savepath = ''
                    else:
                        filesavespl = filesave.split(':/')
                        savepath = '\\\\192.168.120.85\\vpdata\\PROJECT\\'+filesavespl[1].replace('/','\\')+'\\'
                else:
                    savepath = filesave.replace('/','\\')+'\\'

                lineEdit_3.setText(str(savepath))

            except Exception:
                QMessageBox.critical(glowidget,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))


        def anyfunction1(msg1):

            try:

                lineEdit_4.setText(msg1)

            except Exception:
                QMessageBox.critical(glowidget,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def anyfunction2(msg1):

            try:
            
                lineEdit_5.setText(msg1)

            except Exception:
                QMessageBox.critical(glowidget,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def anyfunction3(msg1):

            try:
            
                lineEdit_6.setText(msg1)

            except Exception:
                QMessageBox.critical(glowidget,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def anyfunction4(msg1):

            try:
            
                lineEdit_7.setText(msg1)

            except Exception:
                QMessageBox.critical(glowidget,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def anyfunction5(msg1):

            try:
            
              lineEdit_8.setText(msg1)

            except Exception:
                QMessageBox.critical(glowidget,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))
        
        def anyfunction6(msg1):

            try:
            
              lineEdit_9.setText(msg1)

            except Exception:
                QMessageBox.critical(glowidget,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def dateclick1():

            try:

                global window2
                window2 = secondwindow()
                window2.command1.connect(MyWindow.anyfunction1)

            except Exception:
                QMessageBox.critical(glowidget,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def dateclick2():

            try:
        
                global window2
                window2 = secondwindow()
                window2.command1.connect(MyWindow.anyfunction2)

            except Exception:
                QMessageBox.critical(glowidget,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def dateclick3():

            try:
        
                global window2
                window2 = secondwindow()
                window2.command1.connect(MyWindow.anyfunction3)
                
            except Exception:
                QMessageBox.critical(glowidget,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def dateclick4():

            try:
        
                global window2
                window2 = secondwindow()
                window2.command1.connect(MyWindow.anyfunction4)

            except Exception:
                QMessageBox.critical(glowidget,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def dateclick5():

            try:

                global window2
                window2 = secondwindow()
                window2.command1.connect(MyWindow.anyfunction5)

            except Exception:
                QMessageBox.critical(glowidget,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def dateclick6():

            try:
            
                global window2
                window2 = secondwindow()
                window2.command1.connect(MyWindow.anyfunction6)

            except Exception:
                QMessageBox.critical(glowidget,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def closecheck():

            try:
            
                if lineEdit.text() != '':
                    msgbox = QMessageBox.question(glowidget,'확  인','변경사항이 저장되지 않습니다.\n종료하시겠습니까?')
                    if msgbox == QMessageBox.Yes:
                        glotab.removeTab(glotab.currentIndex())
                    elif msgbox == QMessageBox.No:
                        pass
                else:
                    glotab.removeTab(glotab.currentIndex())

            except Exception:
                QMessageBox.critical(glowidget,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def push7():

            try:
        
                global headerlist

                filename = QFileDialog.getOpenFileName(glowidget,'Open File', filter="EXEL(*.xlsx)")
                
                if filename[0] == '':
                    pass
                    
                else :

                    loadwb = load_workbook(filename[0])
                    loadws = loadwb['정리취합본']

                    headerlist = []

                    for y in range(1,loadws.max_column):
                        headerlist.append(str(loadws.cell(5,y).value).center(len(str(loadws.cell(5,y).value))+2,' '))

                    if ' 제작완료 ' in headerlist:
                        headerlist.remove(' 제작완료 ')
                    if ' 후처리완료 ' in headerlist:
                        headerlist.remove(' 후처리완료 ')
                    if ' 고객납기 ' in headerlist:
                        headerlist.remove(' 고객납기 ')

                    tableWidget.setColumnCount(len(headerlist))   # tableWidget 테이블 위젯 열 갯수를 headerlist 리스트 갯수만큼 설정
                    tableWidget.setHorizontalHeaderLabels(headerlist)   # tableWidget 테이블 위젯 열 헤더에 headerlist 리스트 입력
                    tableWidget.setAutoScroll(True)   # tableWidget 테이블 위젯 내용이 많아질시 자동 스크롤 생성 허용
                    tableWidget.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)   # tableWidget 테이블 위젯 수직 스크롤단위를 픽셀단위로 변경
                    tableWidget.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)   # tableWidget 테이블 위젯 수평 스크롤단위를 픽셀단위로 변경
                    tableWidget.setEditTriggers(QAbstractItemView.EditTriggers(False))   # tableWidget 테이블 위젯 셀 수정 불가하게 설정
                    tableWidget.setSelectionBehavior(QAbstractItemView.SelectRows)   # tableWidget 테이블 위젯 셀 선택시 행 전체 선택 설정
                    tableWidget.setAlternatingRowColors(True)   # tableWidget 테이블 위젯 연속되는 행 색상 다르게 설정
                    tableWidget.setRowCount(loadws.max_row-5)   # tableWidget 테이블 위젯 연속되는 행 색상 다르게 설정

                    headerindex = list(range(1,len(headerlist)+1))
                    exelindex = list(range(6,loadws.max_row+1))

                    for j,x in enumerate(exelindex):
                        for i,y in enumerate(headerindex):
                            if loadws.cell(x,y).value == None:
                                tableWidget.setItem(j, i, QTableWidgetItem(''))
                            else:
                                tableWidget.setItem(j, i, QTableWidgetItem(str(loadws.cell(x,y).value).center(len(str(loadws.cell(x,y).value))+2," ")))
                                tableWidget.item(j,i).setTextAlignment(Qt.AlignCenter)

                    tableWidget.resizeColumnsToContents()
                    tableWidget.resizeRowsToContents()
                    tableWidget.verticalHeader().hide()

                    loadws = loadwb['작업의뢰서 붙여넣기']

                    global dic

                    dic = {}

                    for i in range(10,loadws.max_row+1):
                        if loadws['B'+str(i)].value == None:
                            pass
                        else:
                            dic[loadws['B'+str(i)].value] = (loadws['A'+str(i)].value,
                                                            loadws['C'+str(i)].value,
                                                            loadws['M'+str(i)].value,
                                                            loadws['O'+str(i)].value,
                                                            loadws['P'+str(i)].value
                                                            )

                    keylist = list(sorted(dic.keys(),reverse=False))

                    credate = []
                    company = []
                    enddate = []
                    delidate = []
                    workname = []

                    for i in range(len(keylist)):
                        key = keylist[i]
                        value = dic.get(key)
                        for j in range(len(value)):
                            if j == 0:
                                credate.append(value[j])
                            elif j == 1:
                                company.append(value[j])
                            elif j == 2:
                                enddate.append(value[j])
                            elif j == 3:
                                delidate.append(value[j])
                            elif j == 4:
                                workname.append(value[j])

                    for i,x in enumerate(credate):
                        credate[i] = str(x).split(' ')[0]
                    for i,x in enumerate(enddate):
                        enddate[i] = str(x).split(' ')[0]
                    for i,x in enumerate(delidate):
                        delidate[i] = str(x).split(' ')[0]

                    keylist = ','.join(list(set(keylist)))
                    credate = sorted(list(set(credate)))[0]
                    company = ','.join(list(set(company)))
                    enddate = sorted(list(set(enddate)))[0]
                    delidate = sorted(list(set(delidate)))[0]
                    workname = ','.join(list(set(workname)))

                    lineEdit.setText(keylist)
                    lineEdit_4.setText(credate)
                    lineEdit_2.setText(company)
                    lineEdit_6.setText(enddate)
                    lineEdit_7.setText(enddate)
                    lineEdit_8.setText(delidate)
                    textEdit_2.setText(workname)

                    frelist = []

                    headi = headerlist.index(' 공정 ')

                    for i in range(tableWidget.rowCount()):
                        gong = tableWidget.item(i,headi).text().replace(' ','')

                        if '제관' in gong:
                            frelist.append(tableWidget.item(i,1).text().replace(' ',''))

                        if len(frelist) == 0:
                            lineEdit_9.setText('0000-00-00')
                        else:
                            lineEdit_9.setText(enddate)

            except Exception:
                QMessageBox.critical(glowidget,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

    class secondwindow(QMainWindow,form_secondclass):
            
        command1 = pyqtSignal(str)

        qss = """
            QWidget {
                color: #000000;
                background: #666;
            }
            QWidget#windowTitle {
                # color: #FFFFFF;
                background: #333;
            }
            QWidget#windowTitle QLabel {
                color: #FFFFFF;
                background: #333;
            }
        """

        def __init__(self):

            try:
                super().__init__()
                self.setWindowModality(Qt.ApplicationModal)
                self.setWindowFlags(Qt.FramelessWindowHint)
                self.setStyleSheet(self.qss)
                self.ui = uic.loadUi("resource/ui/calendar.ui",self)
                a = str(pyautogui.position()).strip('Point(x=',).strip(')').split(', y=')
                self.move(int(a[0])-255,int(a[1]))
                self.show()
                self.calendar.clicked.connect(self.sendcommand)

            except Exception:
                QMessageBox.critical(glowidget,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def sendcommand(self):

            try:

                msg1 = self.calendar.selectedDate().toString("yyyy-MM-dd")
                self.command1.emit(msg1)
                self.close()

            except Exception:
                QMessageBox.critical(glowidget,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def keyPressEvent(self,e):

            try:

                if e.key() == Qt.Key_Escape:
                    self.close()

            except Exception:
                QMessageBox.critical(glowidget,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def mousePressEvent(self,Event):
            
            try:

                self.close()

            except Exception:
                QMessageBox.critical(glowidget,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))
        
    if __name__ == '__main__':
        app = QApplication(sys.argv)
        app1 = QApplication(sys.argv)
        myWindow = MyWindow()
        myWindow.show()
        app.exec_()

except Exception:
    err = traceback.format_exc()
    ErrorLog(str(err))