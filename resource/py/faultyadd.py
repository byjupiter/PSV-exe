#-- codingutf-8 --

from PyQt5 import QtGui


def ErrorLog(error: str):
    current_time = time.strftime("%Y.%m.%d/%H:%M:%S", time.localtime(time.time()))
    path = "\\\\192.168.120.85\\vps공유\\00. VP승인 프로그램모음\\13.개발프로그램\\로그\\Log.txt"
    with open(path, "a") as f:
        f.write(f"[{current_time}] - {error}\n")

try:

    import os,glob,sys,traceback,time,pymysql,pyautogui,re
    from PyQt5.QtWidgets import *
    from PyQt5.QtGui import *
    from PyQt5 import uic
    from PyQt5.QtCore import *
    from datetime import datetime
    from dateutil.relativedelta import relativedelta
    from openpyxl import load_workbook

    sys.path.append('resource\\py\\')

    import setting

    ip = setting.setting.setting('nasdb_ip')
    password = setting.setting.setting('nasdb_password')
    
    class faultyadd(object):
        
        def setupUi(widget,tab):

            try:

                global companydic,c_comboBox,c_lineEdit_2,c_tableWidget_2,c_pushButton,c_tableWidget,label,comboBoxList,names,glotab,glowidget,pushButton_4

                glotab = tab
                glowidget = widget

                c_centralwidget = QWidget()
                c_centralwidget.setObjectName("c_centralwidget")
                gridLayout = QGridLayout(c_centralwidget)
                gridLayout.setObjectName("gridLayout")
                c_tableWidget_2 = QTableWidget(c_centralwidget)
                c_tableWidget_2.setEditTriggers(QAbstractItemView.NoEditTriggers)
                c_tableWidget_2.setDragDropMode(QAbstractItemView.NoDragDrop)
                c_tableWidget_2.setSelectionBehavior(QAbstractItemView.SelectRows)
                c_tableWidget_2.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
                c_tableWidget_2.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
                c_tableWidget_2.setObjectName("c_tableWidget_2")
                c_tableWidget_2.setColumnCount(0)
                c_tableWidget_2.setRowCount(0)
                c_tableWidget_2.verticalHeader().setVisible(False)
                gridLayout.addWidget(c_tableWidget_2, 2, 0, 2, 1)
                c_horizontalLayout_2 = QHBoxLayout()
                c_horizontalLayout_2.setSizeConstraint(QLayout.SetDefaultConstraint)
                c_horizontalLayout_2.setObjectName("c_horizontalLayout_2")
                c_label = QLabel(c_centralwidget)
                sizePolicy = QSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
                sizePolicy.setHorizontalStretch(0)
                sizePolicy.setVerticalStretch(0)
                sizePolicy.setHeightForWidth(c_label.sizePolicy().hasHeightForWidth())
                c_label.setSizePolicy(sizePolicy)
                c_label.setObjectName("c_label")
                c_horizontalLayout_2.addWidget(c_label)
                c_comboBox = QComboBox(c_centralwidget)
                sizePolicy = QSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)
                sizePolicy.setHorizontalStretch(0)
                sizePolicy.setVerticalStretch(0)
                sizePolicy.setHeightForWidth(c_comboBox.sizePolicy().hasHeightForWidth())
                c_comboBox.setSizePolicy(sizePolicy)
                c_comboBox.setMinimumSize(QSize(100, 0))
                font = QFont()
                font.setPointSize(12)
                c_comboBox.setFont(font)
                c_comboBox.setObjectName("c_comboBox")
                c_horizontalLayout_2.addWidget(c_comboBox)
                c_label_2 = QLabel(c_centralwidget)
                sizePolicy = QSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
                sizePolicy.setHorizontalStretch(0)
                sizePolicy.setVerticalStretch(0)
                sizePolicy.setHeightForWidth(c_label_2.sizePolicy().hasHeightForWidth())
                c_label_2.setSizePolicy(sizePolicy)
                c_label_2.setObjectName("c_label_2")
                c_horizontalLayout_2.addWidget(c_label_2)
                c_lineEdit_2 = QLineEdit(c_centralwidget)
                c_lineEdit_2.setObjectName("c_lineEdit_2")
                c_horizontalLayout_2.addWidget(c_lineEdit_2)
                c_pushButton_3 = QPushButton(c_centralwidget)
                c_pushButton_3.setObjectName("c_pushButton_3")
                c_horizontalLayout_2.addWidget(c_pushButton_3)
                gridLayout.addLayout(c_horizontalLayout_2, 0, 0, 1, 1)
                c_label_4 = QLabel(c_centralwidget)
                sizePolicy = QSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
                sizePolicy.setHorizontalStretch(0)
                sizePolicy.setVerticalStretch(0)
                sizePolicy.setHeightForWidth(c_label_4.sizePolicy().hasHeightForWidth())
                c_label_4.setSizePolicy(sizePolicy)
                c_label_4.setAlignment(Qt.AlignCenter)
                c_label_4.setObjectName("c_label_4")
                gridLayout.addWidget(c_label_4, 1, 2, 1, 1)
                label = QLabel(c_centralwidget)
                sizePolicy = QSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
                sizePolicy.setHorizontalStretch(0)
                sizePolicy.setVerticalStretch(0)
                sizePolicy.setHeightForWidth(label.sizePolicy().hasHeightForWidth())
                label.setSizePolicy(sizePolicy)
                label.setText("")
                label.setObjectName("label")
                gridLayout.addWidget(label, 4, 0, 1, 1)
                c_label_3 = QLabel(c_centralwidget)
                sizePolicy = QSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
                sizePolicy.setHorizontalStretch(0)
                sizePolicy.setVerticalStretch(0)
                sizePolicy.setHeightForWidth(c_label_3.sizePolicy().hasHeightForWidth())
                c_label_3.setSizePolicy(sizePolicy)
                c_label_3.setAlignment(Qt.AlignCenter)
                c_label_3.setObjectName("c_label_3")
                gridLayout.addWidget(c_label_3, 1, 0, 1, 1)
                horizontalLayout = QHBoxLayout()
                horizontalLayout.setObjectName("horizontalLayout")
                spacerItem = QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)
                horizontalLayout.addItem(spacerItem)
                pushButton_4 = QPushButton(c_centralwidget)
                pushButton_4.setFlat(True)
                pushButton_4.setObjectName("pushButton_4")
                horizontalLayout.addWidget(pushButton_4)
                pushButton = QPushButton(c_centralwidget)
                pushButton.setObjectName("pushButton")
                horizontalLayout.addWidget(pushButton)
                pushButton_2 = QPushButton(c_centralwidget)
                pushButton_2.setObjectName("pushButton_2")
                horizontalLayout.addWidget(pushButton_2)
                gridLayout.addLayout(horizontalLayout, 0, 2, 1, 1)
                pushButton_3 = QPushButton(c_centralwidget)
                pushButton_3.setObjectName("pushButton_3")
                gridLayout.addWidget(pushButton_3, 5, 0, 1, 1)
                c_verticalLayout = QVBoxLayout()
                c_verticalLayout.setSpacing(6)
                c_verticalLayout.setObjectName("c_verticalLayout")
                spacerItem1 = QSpacerItem(20, 40, QSizePolicy.Minimum, QSizePolicy.Expanding)
                c_verticalLayout.addItem(spacerItem1)
                c_pushButton = QPushButton(c_centralwidget)
                c_pushButton.setMinimumSize(QSize(50, 50))
                c_pushButton.setMaximumSize(QSize(50, 50))
                c_pushButton.setObjectName("c_pushButton")
                c_verticalLayout.addWidget(c_pushButton)
                c_pushButton_2 = QPushButton(c_centralwidget)
                c_pushButton_2.setMinimumSize(QSize(50, 50))
                c_pushButton_2.setMaximumSize(QSize(50, 50))
                c_pushButton_2.setObjectName("c_pushButton_2")
                c_verticalLayout.addWidget(c_pushButton_2)
                spacerItem2 = QSpacerItem(20, 40, QSizePolicy.Minimum, QSizePolicy.Expanding)
                c_verticalLayout.addItem(spacerItem2)
                gridLayout.addLayout(c_verticalLayout, 0, 1, 6, 1)
                c_tableWidget = QTableWidget(c_centralwidget)
                c_tableWidget.setMaximumSize(QSize(16777211, 16777215))
                c_tableWidget.setEditTriggers(QAbstractItemView.AllEditTriggers)
                c_tableWidget.setDragDropMode(QAbstractItemView.NoDragDrop)
                c_tableWidget.setSelectionBehavior(QAbstractItemView.SelectRows)
                c_tableWidget.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
                c_tableWidget.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
                c_tableWidget.setObjectName("c_tableWidget")
                c_tableWidget.setColumnCount(0)
                c_tableWidget.setRowCount(0)
                c_tableWidget.verticalHeader().setVisible(False)
                gridLayout.addWidget(c_tableWidget, 2, 2, 4, 1)
                gridLayout.setColumnStretch(0, 1)
                gridLayout.setColumnStretch(2, 2)
                gridLayout.setRowStretch(1, 1)
                gridLayout.setRowStretch(2, 1)
                gridLayout.setRowStretch(3, 1)
                gridLayout.setRowStretch(4, 2)
                widget.setLayout(gridLayout)
                _translate = QCoreApplication.translate
                c_label.setText(_translate("c_MainWindow", "업체명 : "))
                c_label_2.setText(_translate("c_MainWindow", "GEO 이름 : "))
                c_pushButton_3.setText(_translate("c_MainWindow", "검색"))
                c_label_4.setText(_translate("c_MainWindow", "등록 리스트"))
                c_label_3.setText(_translate("c_MainWindow", "검색 리스트"))
                pushButton_4.setText(_translate("c_MainWindow", "요청자 : "))
                pushButton.setText(_translate("c_MainWindow", "저장"))
                pushButton_2.setText(_translate("c_MainWindow", "닫기"))
                pushButton_3.setText(_translate("c_MainWindow", "색인"))
                c_pushButton.setText(_translate("c_MainWindow", "→"))
                c_pushButton_2.setText(_translate("c_MainWindow", "←"))

                faultyname(tab,pushButton_4.text())

                header = [' 관리번호 ',
                          ' 고객사 '
                          ' 도면번호 ',
                          ' 리비전 ',
                          ' 파트번호 ',
                          ' 구성수량 ',
                          ' 재질 ',
                          ' 두께 ',
                          ' 작업수량 ',
                          ' 요청자 ',
                          ' 불량공정 ',
                          ' 불량유형 ',
                          ' 비고 '
                         ]

                c_tableWidget.setColumnCount(len(header))
                c_tableWidget.setHorizontalHeaderLabels(header)
                c_tableWidget.resizeColumnsToContents()
                c_tableWidget.resizeRowsToContents()

                header = [' 도면번호 ',
                        ' 리비전 ',
                        ' 파트번호 ',
                        ' 구성수량 ',
                        ' 재질 ',
                        ' 두께 '
                        ]
                        
                c_tableWidget_2.setColumnCount(len(header))
                c_tableWidget_2.setHorizontalHeaderLabels(header)
                c_tableWidget_2.resizeColumnsToContents()
                c_tableWidget_2.resizeRowsToContents()

                dir_path = "\\\\192.168.120.85\\vpdata\\PROJECT\\"

                companydic = {}

                directories = glob.glob(dir_path+'*')

                for dir in directories:
                    if ' ' in dir and '미거래' not in dir:
                        companydic[str(dir).split(' ')[1]] = dir

                companylist = sorted(companydic.keys())
                c_comboBox.addItems(companylist)

                c_pushButton_3.clicked.connect(faultyadd.search) 
                c_pushButton.clicked.connect(faultyadd.add)
                c_tableWidget_2.currentCellChanged.connect(faultyadd.png)
                pushButton_3.clicked.connect(faultyadd.indexes)

                c_tableWidget.clicked.connect(faultyadd.faultydouble)
                c_tableWidget.itemChanged.connect(faultyadd.resize)
                pushButton_2.clicked.connect(faultyadd.closebtn)
                pushButton.clicked.connect(faultyadd.savebtn)
                pushButton_4.clicked.connect(faultyadd.nameclick)

                faultyadd.diction()

                comboBoxList = []
                names = []

                c_pushButton_2.clicked.connect(faultyadd.delbtn)

            except Exception:
                QMessageBox.critical(QWidget(),'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))
            
        def nameclick():
            
            faultyname(glotab,pushButton_4.text())

        def nameset(text):

            pushButton_4.setText(text)
            row = c_tableWidget.rowCount()
            if row != 0:
                name = text.replace('요청자 : ','')
                for i in range(row):
                    c_tableWidget.setItem(i,9,QTableWidgetItem(name))
                    c_tableWidget.item(i,9).setTextAlignment(Qt.AlignCenter)

        def delbtn():
            
            dellist = []

            for x in c_tableWidget.selectedRanges():

                for ran in range(x.topRow(),x.bottomRow()+1):

                    dellist.append(ran)
                
            dellist.sort(reverse=True)

            for y in dellist:

                c_tableWidget.removeRow(y)

        def savebtn():

            if c_tableWidget.rowCount() != 0:

                msgbox = QMessageBox.question(glowidget,'확  인','저장하시겠습니까?')

                if msgbox == QMessageBox.Yes:

                    
                    for i in range(c_tableWidget.rowCount()):

                        if c_tableWidget.item(i,0) == None or c_tableWidget.item(i,0).text() == '':
    
                            QMessageBox.warning(glowidget,'확  인','관리번호를 입력해주세요.')
                            break

                        elif c_tableWidget.item(i,8) == None or c_tableWidget.item(i,8).text() == '':

                            QMessageBox.warning(glowidget,'확  인','작업수량을 입력해주세요.')
                            break

                        elif c_tableWidget.item(i,10) == None or c_tableWidget.item(i,10).text() == '':

                            QMessageBox.warning(glowidget,'확  인','불량공정을 입력해주세요.')
                            break

                        elif c_tableWidget.item(i,11) == None or c_tableWidget.item(i,11).text() == '':

                            QMessageBox.warning(glowidget,'확  인','불량유형을 입력해주세요.')
                            break
                        
                        else:
                            faultyadd.dbsave(i)
                            QMessageBox.information(glowidget,'저  장','저장되었습니다.')
                            glotab.removeTab(glotab.currentIndex())
      
                elif msgbox == QMessageBox.No:

                    pass

        def dbsave(i):

            conn = pymysql.connect(host=ip, user='user', password=password, db='vps_planner', charset='utf8', port=1980)   # 데이터 베이스 접속 내용을 conn 변수에 저장
            sql = "SELECT MAX(No) FROM faulty"   # 데이터베이스 명령어를 spl 변수에 저장
            with conn.cursor() as cur:
                cur.execute(sql)

            no = int(cur.fetchone()[0])
            today = datetime.strftime(datetime.today(),'%Y-%m-%d')
            
            # for i in range(c_tableWidget.rowCount()):

            sql = "SELECT 판금완료일 FROM planner WHERE 관리번호 = %s"   # 데이터베이스 명령어를 spl 변수에 저장
            with conn.cursor() as cur:
                cur.execute(sql,str(c_tableWidget.item(i,0).text().replace(' ','')))

            date = str(cur.fetchone()[0])
            no += 1
            com = c_tableWidget.item(i,1).text().replace(' ','')
            num = c_tableWidget.item(i,0).text().replace(' ','')
            dwg = c_tableWidget.item(i,2).text().replace(' ','')+"_"+c_tableWidget.item(i,3).text().replace(' ','')
            part = c_tableWidget.item(i,4).text().replace(' ','')
            mt = c_tableWidget.item(i,6).text().replace(' ','')
            tk = float(c_tableWidget.item(i,7).text().replace(' ','').replace('T',''))
            ea = int(c_tableWidget.item(i,8).text().replace(' ',''))
            name = c_tableWidget.item(i,9).text().replace(' ','')
            gong = c_tableWidget.item(i,10).text().replace(' ','')
            yu = c_tableWidget.item(i,11).text().replace(' ','')
            oth = c_tableWidget.item(i,12).text().replace(' ','')

            sql = "INSERT INTO faulty VALUE (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"   # 데이터베이스 명령어를 spl 변수에 저장
            with conn.cursor() as cur:
                cur.execute(sql,(no,'처리전',today,com,date,num,dwg,part,mt,tk,ea,name,gong,yu,oth))
                conn.commit()
            
            conn.close()

        def closebtn():

            if c_tableWidget.rowCount() != 0:
                msgbox = QMessageBox.question(glowidget,'확  인','변경사항이 저장되지 않습니다.\n종료하시겠습니까?')
                if msgbox == QMessageBox.Yes:
                    glotab.removeTab(glotab.currentIndex())
                elif msgbox == QMessageBox.No:
                    pass
            else:
                glotab.removeTab(glotab.currentIndex())

        def resize():
             
            # and c_tableWidget.item(c_tableWidget.currentRow(),3).text() == '상세정보'
            if c_tableWidget.currentIndex().column() == 7 and ' EA' not in c_tableWidget.currentItem().text() :
                cell = c_tableWidget.item(c_tableWidget.currentIndex().row(),c_tableWidget.currentIndex().column()).text()
                recell = re.sub(r'[^0-9]', '', cell) + ' EA'
                c_tableWidget.setItem(c_tableWidget.currentIndex().row(),7,QTableWidgetItem(recell))
                # print(c_tableWidget.rowCount())
                # c_tableWidget.setRowHidden(0,False)
    
            if c_tableWidget.currentItem() != None:
                c_tableWidget.currentItem().setTextAlignment(Qt.AlignCenter)
                c_tableWidget.resizeColumnsToContents()

        def png():

            if c_tableWidget_2.item(c_tableWidget_2.currentRow(),0) != None:

                name = c_tableWidget_2.item(c_tableWidget_2.currentRow(),0).text().strip()
                rev = c_tableWidget_2.item(c_tableWidget_2.currentRow(),1).text().strip()
                part = c_tableWidget_2.item(c_tableWidget_2.currentRow(),2).text().strip()
                set = c_tableWidget_2.item(c_tableWidget_2.currentRow(),3).text().strip()

                keylist = [name,rev,part,set]

                keys = '_'.join(keylist)

                if keys in dic:
                    pixmap = QPixmap(dic[keys]).scaledToHeight(label.height())

                    if pixmap.height() > pixmap.width():
                        pixmap = pixmap.scaledToHeight(label.height())
                        if pixmap.width() > label.width():
                            pixmap = pixmap.scaledToWidth(label.width())

                    elif pixmap.width() > pixmap.height():
                        pixmap = pixmap.scaledToWidth(label.width())
                        if pixmap.height() > label.height():
                            pixmap = pixmap.scaledToHeight(label.height())

                    label.setPixmap(pixmap)
                    label.setAlignment(Qt.AlignCenter)
                else:
                    pixmap = QPixmap('resource\\image\\noimage.png').scaledToWidth(label.width())
                    label.setPixmap(pixmap)
                    label.setAlignment(Qt.AlignCenter)

            else:
                pass

        def diction():

            global dic

            txtpath = "\\\\192.168.120.85\\vps공유\\00. VP승인 프로그램모음\\13.개발프로그램\\로그\\indexes"

            f = open(txtpath,'r',encoding='UTF8')
            lines = f.read()
            f.close()

            dic = {}
            data_list = lines.split("\n")

            for data in data_list:
                pair = data.split(":")
                if pair[0] == '':
                    pass
                else:
                    keys = pair[0].strip()
                    values = pair[1].strip()
                    dic[keys] = values

        def indexes():

            message = QMessageBox.question(QWidget(), '색인 진행', '색인에는 약 1분정도 소요됩니다. \n 진행하시겠습니까?', 
                                 QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            
            if message == QMessageBox.Yes:

                global dic

                path = '\\\\192.168.120.60\\Archiv\\Artikel'

                dic = {}

                for i,(root, directories, files) in enumerate(os.walk(path)):
                    if 'DrawingFile' not in root:
                        for file in files:
                            if 'GEO' in file and 'JIG' not in file and 'TEST' not in file:
                                file = file.split('_')
                                del file[-1]
                                file = '_'.join(file)
                                keyed = file
                            if 'FabGenerated' in file:
                                valued = str(root+'\\'+file)
                                dic[keyed] = valued
                            else:
                                pass

                txtpath = "\\\\192.168.120.85\\vps공유\\00. VP승인 프로그램모음\\13.개발프로그램\\로그\\indexes"
                with open(txtpath, "w",encoding='UTF8') as i:
                    for key,value in dic.items():
                        i.write(f'{key} : {value}\n')

                faultyadd.diction()

                QMessageBox.information(QWidget(),'색인 완료','색인이 완료되었습니다.')

            else:

                QMessageBox.information(QWidget(),'색인 취소','색인이 취소되었습니다.')

        def faultydouble():

            try:
                if c_tableWidget.currentColumn() == 0:
                    faultyinsert1()
                elif c_tableWidget.currentColumn() == 10 or c_tableWidget.currentColumn() == 11:
                    faultyinsert2()

            except Exception:
                QMessageBox.critical(QWidget(),'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def add():

            try:

                header = [' 관리번호 ',
                          ' 고객사 ',
                          ' 도면번호 ',
                          ' 리비전 ',
                          ' 파트번호 ',
                          ' 구성수량 ',
                          ' 재질 ',
                          ' 두께 ',
                          ' 작업수량 ',
                          ' 요청자 ',
                          ' 불량공정 ',
                          ' 불량유형 ',
                          ' 비고 '
                         ]

                c_tableWidget.setColumnCount(len(header))
                c_tableWidget.setHorizontalHeaderLabels(header)

                global baserow,endrow

                baserow = c_tableWidget.rowCount()
                name = pushButton_4.text().replace('요청자 : ','')

                for x in c_tableWidget_2.selectedRanges():

                    for ran in range(x.topRow(),x.bottomRow()+1):

                        i = c_tableWidget.rowCount()
                        c_tableWidget.insertRow(i)
                        c_tableWidget.setItem(i,0,QTableWidgetItem(''))
                        c_tableWidget.item(i,0).setFlags(c_tableWidget.item(i,0).flags() ^  Qt.ItemFlag.ItemIsEditable)
                        company = c_comboBox.currentText()
                        c_tableWidget.setItem(i,1,QTableWidgetItem(company))
                        c_tableWidget.item(i,1).setTextAlignment(Qt.AlignCenter)
                        c_tableWidget.item(i,1).setFlags(c_tableWidget.item(i,1).flags() ^  Qt.ItemFlag.ItemIsEditable)

                        for ii in range(9,12):

                            if ii == 9:
                                c_tableWidget.setItem(i,ii,QTableWidgetItem(name))
                                c_tableWidget.item(i,ii).setTextAlignment(Qt.AlignCenter)
                            elif ii != 9:
                                c_tableWidget.setItem(i,ii,QTableWidgetItem(''))
                            c_tableWidget.item(i,ii).setFlags(c_tableWidget.item(i,ii).flags() ^  Qt.ItemFlag.ItemIsEditable)

                        items = []

                        for j in range(6):

                            item = c_tableWidget_2.item(ran,j).text()
                            c_tableWidget.setItem(i,j+2,QTableWidgetItem(item))
                            c_tableWidget.item(i,j+2).setTextAlignment(Qt.AlignCenter)
                            c_tableWidget.item(i,j+2).setFlags(c_tableWidget.item(i,j+2).flags() ^  Qt.ItemFlag.ItemIsEditable)
                            items.append(item.replace(' ',''))

                            if j == 3:

                                itemed = '_'.join(items)

                        if itemed not in names:

                            names.append(itemed)
                            
                        else:

                            c_tableWidget.removeRow(c_tableWidget.rowCount()-1)
                            QMessageBox.warning(QWidget(),'중복 발생','중복되는 제품이 있습니다.\n확인해주세요.\n'+itemed)
                        

                endrow = c_tableWidget.rowCount()

                c_tableWidget.resizeColumnsToContents()
                c_tableWidget.resizeRowsToContents()

                name = list(set(names))

                # print(names)
                # print(name)

            except Exception:
                QMessageBox.critical(QWidget(),'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def search():

            try:

                c_tableWidget_2.clear()
                c_tableWidget_2.setRowCount(0)

                com_path = companydic[c_comboBox.currentText()]

                geo = {}
                name = c_lineEdit_2.text()

                for (root, directories, files) in os.walk(com_path):
                    for i,file in enumerate(files):
                        if name in file and 'JIG' not in file and '.GEO' in file and 'TEST' not in file:
                            file_path = os.path.join(root, file)
                            if len(file_path.split('\\')[-1].split('_')) == 5 and 'JIG' not in file_path.split('\\')[-2] and 'GEO' not in file_path.split('\\')[-2]:
                                if '환봉' not in file_path.split('\\')[-2] and ' ' in file_path.split('\\')[-2]:

                                    named = file_path.split('\\')[-1].split('_')[0]
                                    revision = file_path.split('\\')[-1].split('_')[1]
                                    partnum = file_path.split('\\')[-1].split('_')[2]
                                    setea = file_path.split('\\')[-1].split('_')[3]
                                    if len(file_path.split('\\')[-2].split(' ')) == 3:
                                        mt = file_path.split('\\')[-2].split(' ')[0]+file_path.split('\\')[-2].split(' ')[2]
                                    elif len(file_path.split('\\')[-2].split(' ')) == 2:
                                        mt = file_path.split('\\')[-2].split(' ')[0]
                                    tk = file_path.split('\\')[-2].split(' ')[1]

                                    geo[file] = [named,revision,partnum,setea,mt,tk,file_path]

                            else:
                                pass

                geo = dict(sorted(geo.items()))
                partnumed = geo.keys()
                
                header = [' 도면번호 ',
                        ' 리비전 ',
                        ' 파트번호 ',
                        ' 구성수량 ',
                        ' 재질 ',
                        ' 두께 '
                        ]

                c_tableWidget_2.setColumnCount(len(header))
                c_tableWidget_2.setHorizontalHeaderLabels(header)

                for i,part in enumerate(partnumed):
                    c_tableWidget_2.insertRow(c_tableWidget_2.rowCount())
                    for j in range(len(header)):
                        c_tableWidget_2.setItem(i,j,QTableWidgetItem(geo[part][j].center(len(geo[part][j])+2,' ')))
                        c_tableWidget_2.item(i,j).setTextAlignment(Qt.AlignCenter)
                    
                c_tableWidget_2.resizeColumnsToContents()
                c_tableWidget_2.resizeRowsToContents()

            except Exception:
                QMessageBox.critical(QWidget(),'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        
        def setnumber(number):

            c_tableWidget.setItem(c_tableWidget.currentRow(),0,QTableWidgetItem(str(number).center(len(str(number))+2,' ')))
            c_tableWidget.item(c_tableWidget.currentRow(),0).setFlags(c_tableWidget.item(c_tableWidget.currentRow(),0).flags() ^  Qt.ItemFlag.ItemIsEditable)
            c_tableWidget.item(c_tableWidget.currentRow(),0).setTextAlignment(Qt.AlignCenter)
            faultyadd.resize()

        def setprowhat(pro,what):
    
            c_tableWidget.setItem(c_tableWidget.currentRow(),10,QTableWidgetItem(str(pro).center(len(str(pro))+2,' ')))
            c_tableWidget.item(c_tableWidget.currentRow(),10).setFlags(c_tableWidget.item(c_tableWidget.currentRow(),10).flags() ^  Qt.ItemFlag.ItemIsEditable)
            c_tableWidget.item(c_tableWidget.currentRow(),10).setTextAlignment(Qt.AlignCenter)
            c_tableWidget.setItem(c_tableWidget.currentRow(),11,QTableWidgetItem(str(what).center(len(str(what))+2,' ')))
            c_tableWidget.item(c_tableWidget.currentRow(),11).setFlags(c_tableWidget.item(c_tableWidget.currentRow(),11).flags() ^  Qt.ItemFlag.ItemIsEditable)
            c_tableWidget.item(c_tableWidget.currentRow(),11).setTextAlignment(Qt.AlignCenter)
            faultyadd.resize()

    form_faultyinsert1 = uic.loadUiType("resource/ui/faultyinsert1.ui")[0]   # UI 파일 불러와서 변수에 저장

    class faultyinsert1(QMainWindow,form_faultyinsert1):
            
        def __init__(self):
    
            try:

                super().__init__()
                self.setWindowModality(Qt.ApplicationModal)
                self.ui = uic.loadUi("resource/ui/faultyinsert1.ui",self)
                self.show()
                self.setWindowModality(Qt.ApplicationModal)

                today = datetime.today()
                month_3 = relativedelta(months=3)
                preday = today-month_3
                tomoday = today+month_3

                self.lineEdit_2.setText(tomoday.strftime('%Y-%m-%d'))
                self.lineEdit.setText(preday.strftime('%Y-%m-%d'))

                self.insert1table()

                self.pushButton_3.clicked.connect(self.dateclick1)
                self.pushButton_4.clicked.connect(self.dateclick2)
                self.pushButton_5.clicked.connect(self.insert1table)
                self.pushButton.clicked.connect(self.next1)


            except Exception:
                QMessageBox.critical(QWidget(),'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def next1(self):

            numbers = []
            for i in range(self.tableWidget.rowCount()):

                ckbox = self.tableWidget.cellWidget(i, 0).findChildren(QCheckBox)

                if ckbox[0].isChecked() == True:

                    numbers.append(self.tableWidget.item(i,1).text().replace(' ',''))

            if len(numbers) != 0:

                if len(numbers) > 1 and '확인불가' in numbers:

                    QMessageBox.critical(QWidget(),'에러 발생','관리번호 선택오류!')
                
                else:

                    number = ' / '.join(numbers)  
                    faultyadd.setnumber(number)
                    self.close()
                    faultyinsert2()

            else:

                QMessageBox.critical(QWidget(),'에러 발생','관리번호를 선택해주세요.')

        def insert1table(self):
        
            try:

                name = '%'+c_tableWidget.item(c_tableWidget.currentRow(),1).text()+'%'
                conn = pymysql.connect(host=ip, user='user', password=password, db='vps_planner', charset='utf8', port=1980)   # 데이터 베이스 접속 내용을 conn 변수에 저장
                sql = "SELECT 관리번호,판금완료일 FROM sidedater WHERE 테이블내용 LIKE %s AND 판금완료일 BETWEEN %s AND %s"   # 데이터베이스 명령어를 spl 변수에 저장
                with conn.cursor() as cur:
                    cur.execute(sql,(name,self.lineEdit.text(),self.lineEdit_2.text()))
                    allnum = cur.fetchall()
                    allnum = list(set(allnum))
                    allnum.sort()
                conn.close()

                row = c_tableWidget.currentRow()
                if row != 0:
                    item = c_tableWidget.item(row-1,0).text()
                    if row != 0:
                        if item != '' and '/' in item:
                            item = item.replace(' ', '').split('/')
                        else:
                            item = item.replace(' ', '')
                else:
                    item = ''

                self.tableWidget.setColumnCount(3)
                self.tableWidget.setHorizontalHeaderLabels([' 선택 ',' 관리번호 ',' 판금완료일 '])
                self.tableWidget.setRowCount(1)
                self.tableWidget.setItem(0,1,QTableWidgetItem(' 확인불가 '))
                self.tableWidget.item(0,1).setTextAlignment(Qt.AlignCenter)
                ckbox = QCheckBox()
                if '확인불가' in item:
                    ckbox.setChecked(True)
                cellWidget = QWidget()
                layoutCB = QHBoxLayout(cellWidget)
                layoutCB.addWidget(ckbox)
                layoutCB.setAlignment(Qt.AlignCenter)
                layoutCB.setContentsMargins(0,0,0,0)
                cellWidget.setLayout(layoutCB)
                self.tableWidget.setCellWidget(0,0,cellWidget)

                for i,nums in enumerate(allnum):
                    self.tableWidget.insertRow(i+1)
                    for j,num in enumerate(nums):
                        self.tableWidget.setItem(i+1,j+1,QTableWidgetItem(str(num).center(len(str(num))+2,' ')))
                        ckbox = QCheckBox()     
                        if str(nums[0]) in item:
                            ckbox.setChecked(True)
                        cellWidget = QWidget()
                        layoutCB = QHBoxLayout(cellWidget)
                        layoutCB.addWidget(ckbox)
                        layoutCB.setAlignment(Qt.AlignCenter)            
                        layoutCB.setContentsMargins(0,0,0,0)
                        cellWidget.setLayout(layoutCB)        
                        self.tableWidget.setCellWidget(i+1,0,cellWidget)

                self.tableWidget.resizeColumnsToContents()
                self.tableWidget.resizeRowsToContents()

            except Exception:
                QMessageBox.critical(self,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def dateclick1(self):
        
            try:

                window2 = secondwindow(datetime.strptime(self.lineEdit.text().replace(' ',''),'%Y-%m-%d'))
                window2.command1.connect(self.anyfunction1)

            except Exception:
                QMessageBox.critical(self,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def anyfunction1(self,msg1):

            try:

                self.lineEdit.setText(msg1)

            except Exception:
                QMessageBox.critical(self,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def dateclick2(self):
        
            try:

                window2 = secondwindow(datetime.strptime(self.lineEdit_2.text().replace(' ',''),'%Y-%m-%d'))
                window2.command1.connect(self.anyfunction2)

            except Exception:
                QMessageBox.critical(self,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))


        def anyfunction2(self,msg1):

            try:

                self.lineEdit_2.setText(msg1)

            except Exception:
                QMessageBox.critical(self,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

    form_secondclass = uic.loadUiType("resource/ui/calendar.ui")[0]   # UI 파일 불러와서 변수에 저장

    class secondwindow(QMainWindow,form_secondclass):
            
        command1 = pyqtSignal(str)

        qss = """
            QWidget {
                color: #000000;
                background: #666;
            }
            QWidget#windowTitle {
                color: #FFFFFF;
                background: #333;
            }
            QWidget#windowTitle QLabel {
                color: #FFFFFF;
                background: #333;
            }
        """

        def __init__(self,date):

            try:
                super().__init__()
                self.setWindowModality(Qt.ApplicationModal)
                self.setWindowFlags(Qt.FramelessWindowHint)
                self.setStyleSheet(self.qss)
                self.ui = uic.loadUi("resource/ui/calendar.ui",self)
                a = str(pyautogui.position()).strip('Point(x=',).strip(')').split(', y=')
                self.move(int(a[0])-255,int(a[1]))
                self.show()
                self.calendar.clicked.connect(self.sendcommand)
                self.calendar.setSelectedDate(date)

            except Exception:
                QMessageBox.critical(QWidget(),'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                # ErrorLog(str(err))

        def sendcommand(self):

            try:

                msg1 = self.calendar.selectedDate().toString("yyyy-MM-dd")
                self.command1.emit(msg1)
                self.close()

            except Exception:
                QMessageBox.critical(QWidget(),'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                # ErrorLog(str(err))

        def keyPressEvent(self,e):

            try:

                if e.key() == Qt.Key_Escape:
                    self.close()

            except Exception:
                QMessageBox.critical(QWidget(),'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                # ErrorLog(str(err))

        def mousePressEvent(self,Event):
            
            try:

                self.close()

            except Exception:
                QMessageBox.critical(QWidget(),'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                # ErrorLog(str(err))

    form_faultyinsert2 = uic.loadUiType("resource/ui/faultyinsert2.ui")[0]   # UI 파일 불러와서 변수에 저장

    class faultyinsert2(QMainWindow,form_faultyinsert2):
            
        def __init__(self):
    
            try:

                super().__init__()
                self.ui = uic.loadUi("resource/ui/faultyinsert2.ui",self)
                self.show()
                self.setWindowModality(Qt.ApplicationModal)
                self.listtable1()
                self.pushButton.clicked.connect(self.btnclick)

            except Exception:
                QMessageBox.critical(QWidget(),'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def btnclick(self):
            
            if self.listWidget.currentItem() != None:
                pro = self.listWidget.currentItem().text()
            else:
                pro = ''
            if self.listWidget_2.currentItem() != None:
                what = self.listWidget_2.currentItem().text()
            else:
                what = ''

            faultyadd.setprowhat(pro,what)

            self.close()

            selecrow = c_tableWidget.currentRow()
            row = c_tableWidget.rowCount()
            item = c_tableWidget.item(selecrow+1,0)

            if selecrow != row and item != None:
                c_tableWidget.setCurrentCell(selecrow+1,0)
                faultyinsert1()

        def listtable1(self):
        
            try:

                conn = pymysql.connect(host=ip, user='user', password=password, db='vps_planner', charset='utf8', port=1980)   # 데이터 베이스 접속 내용을 conn 변수에 저장
                sql = "SELECT 불량공정 FROM faultysetting"   # 데이터베이스 명령어를 spl 변수에 저장
                with conn.cursor() as cur:
                    cur.execute(sql)
                    pro = cur.fetchall()
                conn.close()
                pro = list(set(pro))
                pro.sort()

                for x in pro:
                    self.listWidget.addItem(x[0])

                row = c_tableWidget.currentRow()

                if row != 0:
                    item = c_tableWidget.item(row-1,10).text().replace(' ','')
                else:
                    item = ''
                
                if item != '':

                    if c_tableWidget.item(row,10).text().replace(' ','') == '':
                    
                        for i in range(self.listWidget.count()):

                            items = self.listWidget.item(i).text()

                            if item == items:

                                self.listWidget.setCurrentRow(i)
                                self.proclick()

                    else:

                        item = c_tableWidget.item(row,10).text().replace(' ','')

                        for i in range(self.listWidget.count()):

                            items = self.listWidget.item(i).text()

                            if item == items:

                                self.listWidget.setCurrentRow(i)
                                self.proclick()

                self.listWidget.currentItemChanged.connect(self.proclick)

            except Exception:
                QMessageBox.critical(self,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def proclick(self):

            try:

                self.listWidget_2.clear()

                conn = pymysql.connect(host=ip, user='user', password=password, db='vps_planner', charset='utf8', port=1980)   # 데이터 베이스 접속 내용을 conn 변수에 저장
                sql = "SELECT 불량유형 FROM faultysetting WHERE 불량공정 = %s"   # 데이터베이스 명령어를 spl 변수에 저장
                with conn.cursor() as cur:
                    cur.execute(sql,(self.listWidget.currentItem().text()))
                    what = cur.fetchall()
                conn.close()
                what = list(set(what))
                what.sort()

                for x in what:

                    self.listWidget_2.addItem(x[0])
                
                row = c_tableWidget.currentRow()
                if row != 0:
                    item = c_tableWidget.item(row-1,11).text().replace(' ','')
                else:
                    item = ''
                
                if item != '':

                    if c_tableWidget.item(row,11).text().replace(' ','') == '':
                        
                        for i in range(self.listWidget_2.count()):

                            items = self.listWidget_2.item(i).text()

                            if item == items:

                                self.listWidget_2.setCurrentRow(i)

                    else:

                        item = c_tableWidget.item(row,11).text().replace(' ','')

                        for i in range(self.listWidget_2.count()):

                            items = self.listWidget_2.item(i).text()

                            if item == items:

                                self.listWidget_2.setCurrentRow(i)

            except Exception:
                QMessageBox.critical(self,'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

    form_faultyname = uic.loadUiType("resource/ui/faultyname1.ui")[0]   # UI 파일 불러와서 변수에 저장

    class faultyname(QMainWindow,form_faultyname):
            
        def __init__(self,tab,name):
    
            try:
                
                self.tab = tab
                self.num = 0
                self.name = name.replace('요청자 : ','')

                super().__init__()
                self.setWindowModality(Qt.ApplicationModal)
                self.ui = uic.loadUi("resource/ui/faultyname1.ui",self)
                self.show()
                self.setWindowModality(Qt.ApplicationModal)
                self.nametable()
                self.lineEdit.setFocus()
                self.setname()
                self.tableWidget.setStyleSheet(
                                    "QTableView::item:selected"
                                    "{"
                                    "background-color : #2DA0FF;"
                                    "selection-color : #000000;"
                                    "}"
                                    )

                self.pushButton.clicked.connect(self.click)
                self.tableWidget.currentCellChanged.connect(self.tableclick)

                if self.name != '':
                    
                    item = self.tableWidget.findItems(' '+self.name+' ',Qt.MatchFlag.MatchExactly)
                    self.tableWidget.setCurrentItem(item[0])

            except Exception:
                QMessageBox.critical(QWidget(),'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
                err = traceback.format_exc()
                ErrorLog(str(err))

        def tableclick(self):

            row = self.tableWidget.currentRow()
            item = self.tableWidget.item(row,2).text().replace(' ','')
            self.lineEdit.setText(item)

        def setname(self):
            
            self.lineEdit.setText(self.name)

        def click(self):

            self.num += 1
            self.close()

        def nametable(self):

            dir = '\\\\192.168.120.85\\vps공유\\00. 사내 연락처 및 카플\\사내연락처'
            files = os.listdir(dir)
            filelist = []

            for file in files : 

                fullname_file = os.path.join(dir, file)

                if '.xlsx' in file and '~$' not in file:

                    filelist.append(fullname_file)

            if len(filelist) == 1:
                
                loadwb = load_workbook(fullname_file)
                loadws = loadwb['Sheet1']
                header = ['부서','파트','성명']
                j = 0
                
                self.tableWidget.setColumnCount(3)
                self.tableWidget.setHorizontalHeaderLabels(header)

                for x in range(3,loadws.max_row):

                    if loadws.cell(x,2).value != None:

                        j = 0
                        department = str(loadws.cell(x,2).value)

                    elif loadws.cell(x,2).value == None:

                        j += 1
                        department = str(loadws.cell(x-j,2).value)
                        
                    if department == '품질출하팀' or department == '제조팀':
                        
                        i = self.tableWidget.rowCount()
                        
                        if loadws.cell(x,3).value != None:
                            
                            part = str(loadws.cell(x,3).value)
                            
                        elif loadws.cell(x,3).value == None:
                            
                            part = str(loadws.cell(x,7).value)
                            
                        name = str(loadws.cell(x,4).value)
                        
                        self.tableWidget.insertRow(i)
                        self.tableWidget.setItem(i, 0, QTableWidgetItem(department.center(len(department)+2," ")))
                        self.tableWidget.setItem(i, 1, QTableWidgetItem(part.center(len(part)+2," ")))
                        self.tableWidget.setItem(i, 2, QTableWidgetItem(name.center(len(name)+2," ")))

            self.tableWidget.resizeColumnsToContents()
            self.tableWidget.resizeRowsToContents()

        def closeEvent(self, event):

            if self.lineEdit.text() == '':

                self.tab.removeTab(0)

            elif self.lineEdit.text() != '' and self.num == 0:

                self.tab.removeTab(0)
            
            elif self.lineEdit.text() != '' and self.num == 1:

                faultyadd.nameset('요청자 : ' + self.lineEdit.text())

except Exception:
    QMessageBox.critical(QWidget(),'에러 발생','에러가 발생했습니다.\n로그를 확인해주세요.')
    err = traceback.format_exc()
    ErrorLog(str(err))