from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5 import uic
from dateutil.relativedelta import relativedelta
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
import pymysql,datetime,sys,time

sys.path.append('resource\\py\\')

import setting
 
form_insertclass = uic.loadUiType("resource/ui/insert.ui")[0]   # UI 파일 불러와서 변수에 저장

class qclist(QMainWindow):   # 메인윈도우 클래스 시작

    def qclisttable(widget,tab):   # plantabwidget 함수 시작
        
        global table1,glotab,gridLayout,lineEdit_4,line2,line3,glowidget
        
        glotab = tab
        glowidget = widget
        
        gridLayout = QGridLayout()
        gridLayout.setObjectName("gridLayout")
        hbox2 = QHBoxLayout()
        hbox2.setObjectName("hbox2")
        labels3 = QLabel()
        labels3.setObjectName("labels3")
        hbox2.addWidget(labels3)
        line2 = QLineEdit()
        line2.setMaximumSize(QSize(110, 20))
        line2.setAlignment(Qt.AlignCenter)
        line2.setObjectName("line2")
        hbox2.addWidget(line2)
        datebtn1 = QPushButton()
        datebtn1.setMinimumSize(QSize(23, 23))
        datebtn1.setMaximumSize(QSize(23, 23))
        datebtn1.setObjectName("datebtn1")
        hbox2.addWidget(datebtn1)
        labels4 = QLabel()
        labels4.setObjectName("labels4")
        hbox2.addWidget(labels4)
        line3 = QLineEdit()
        line3.setMaximumSize(QSize(110, 20))
        line3.setAlignment(Qt.AlignCenter)
        line3.setObjectName("line3")
        hbox2.addWidget(line3)
        datebtn2 = QPushButton()
        datebtn2.setMinimumSize(QSize(23, 23))
        datebtn2.setMaximumSize(QSize(23, 23))
        datebtn2.setObjectName("datebtn2")
        hbox2.addWidget(datebtn2)
        spacerItem = QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)
        hbox2.addItem(spacerItem)
        label_5 = QLabel()
        label_5.setObjectName("label_5")
        hbox2.addWidget(label_5)
        lineEdit_4 = QLineEdit()
        lineEdit_4.setMaximumSize(QSize(200, 20))
        lineEdit_4.setObjectName("lineEdit_4")
        hbox2.addWidget(lineEdit_4)
        pushButton_7 = QPushButton()
        pushButton_7.setObjectName("pushButton_7")
        hbox2.addWidget(pushButton_7)
        refreshbtn = QPushButton()
        refreshbtn.setObjectName("refreshbtn")
        hbox2.addWidget(refreshbtn)
        pushButton = QPushButton()
        pushButton.setObjectName("pushButton")
        hbox2.addWidget(pushButton)
        gridLayout.addLayout(hbox2, 0, 0, 1, 1)
        table1 = QTableWidget()
        table1.setObjectName("table1")
        table1.setColumnCount(0)
        table1.setRowCount(0)
        gridLayout.addWidget(table1, 2, 0, 1, 1)
        hbox2_2 = QHBoxLayout()
        hbox2_2.setObjectName("hbox2_2")
        label = QLabel()
        label.setObjectName("label")
        hbox2_2.addWidget(label)
        spacerItem1 = QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)
        hbox2_2.addItem(spacerItem1)
        pushButton_3 = QPushButton()
        pushButton_3.setObjectName("pushButton_3")
        hbox2_2.addWidget(pushButton_3)
        pushButton_2 = QPushButton()
        pushButton_2.setObjectName("pushButton_2")
        hbox2_2.addWidget(pushButton_2)
        gridLayout.addLayout(hbox2_2, 1, 0, 1, 1)

        widget.setLayout(gridLayout)
        
        _translate = QCoreApplication.translate
        labels3.setText(_translate("grid", "수주일자 : "))
        line2.setInputMask(_translate("grid", "0000 - 00 - 00"))
        datebtn1.setText(_translate("grid", "▼"))
        labels4.setText(_translate("grid", "~"))
        line3.setInputMask(_translate("grid", "0000 - 00 - 00"))
        datebtn2.setText(_translate("grid", "▼"))
        label_5.setText(_translate("grid", "도면번호 검색 : "))
        pushButton_7.setText(_translate("grid", "검색"))
        refreshbtn.setText(_translate("grid", "새로고침"))
        pushButton.setText(_translate("grid", "엑셀저장"))
        table1.setSortingEnabled(True)
        label.setText(_translate("grid", "필터 현황 : "))
        pushButton_3.setText(_translate("grid", "필터설정"))
        pushButton_2.setText(_translate("grid", "필터초기화"))
        
        table1.setSortingEnabled(True)
        
        now = datetime.datetime.now()
        nowstr = now.strftime('%Y-%m-%d')
        qc_premonth = relativedelta(months=int(setting.setting.setting('qc_premonth')))
        start_day_datetime = now - qc_premonth
        start_day = start_day_datetime.strftime('%Y-%m-%d')
        
        line2.setText(start_day)
        line3.setText(nowstr)
        
        qclist.plantable('')
        
        refreshbtn.clicked.connect(lambda: qclist.plantable(''))
        pushButton_7.clicked.connect(qclist.search)
        pushButton.clicked.connect(qclist.excel)
        
    def excel():
        
        wb = Workbook()
        ws = wb.active  

        for i in range(table1.rowCount()):
            
            ii = i + 2
            pasj = 0
            
            for j in range(table1.columnCount()):
                
                if j == 0 or j == 2 or j == 9:
                    
                    pasj += 1
                    pass
                
                else:
                    
                    jj = j + 2 - pasj
                    
                    if table1.item(i,j) == None:

                        ws.cell(ii,jj,'').alignment = Alignment(horizontal='center', vertical='center')
                    
                    else:
                    
                        ws.cell(ii,jj,str(table1.item(i,j).text()).strip().replace("   "," ").replace("  "," ")).alignment = Alignment(horizontal='center', vertical='center')

        path = QFileDialog.getSaveFileName(glowidget,'Save File','','Exel(*.xlsx)')

        try:
            wb.save(path[0])
            QMessageBox.information(glowidget,'저장 완료','저장이 완료되었습니다.')
        except:
            QMessageBox.warning(glowidget,'경로 선택','경로 및 파일이름을 설정해주세요.')
            pass
    
    def changeitem(item):
        
        row = table1.currentRow()

        if item.column() == 18 and item.text() != '':
            
            conn2 = pymysql.connect(host='192.168.120.85', user='user', password='VPsystem1234!!', db='vps_planner', charset='utf8', port=1980)   # 데이터 베이스 접속 내용을 conn 변수에 저장
            bom_id = table1.item(row, 0).text()
            order_no = table1.item(row, 2).text()
            part_no = table1.item(row, 5).text()
            dataed = table1.item(row, 15).text()
            partner_company = table1.item(row,18).text()
            division = table1.item(row, 17).text()
            now = datetime.datetime.now()
            take_out_time = now.strftime('%Y-%m-%d %H:%M:%S')
            data = take_out_time.split(' ')[0]
            
            table1.item(row, 19).setText(data)

            sql = "SELECT qc_time FROM qclist WHERE bom_id = %s AND order_no = %s AND part_no = %s"  # 데이터베이스 명령어를 spl 변수에 저장
            with conn2.cursor() as cur:
                cur.execute(sql,(bom_id,order_no,part_no))
                sqlnowstr = cur.fetchone()
                    
            if sqlnowstr != None:

                sql = "UPDATE qclist SET partner_company = %s,take_out_date = %s WHERE bom_id = %s AND order_no = %s AND part_no = %s"   # 데이터베이스 명령어를 spl 변수에 저장
                with conn2.cursor() as cur:
                    cur.execute(sql,(partner_company,take_out_time,bom_id,order_no,part_no))
                    conn2.commit()

            elif sqlnowstr == None:

                dataed = None
                nowstr = None
                division = None

                sql = "INSERT INTO qclist (bom_id,order_no,part_no,state,qc_time,division,partner_company,take_out_date) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)"   # 데이터베이스 명령어를 spl 변수에 저장
                with conn2.cursor() as cur:
                    cur.execute(sql,(bom_id,order_no,part_no,dataed,nowstr,division,partner_company,take_out_time))
                    conn2.commit()
                print(dataed,1)

        if item.column() == 17 and item.text() != '':

            conn2 = pymysql.connect(host='192.168.120.85', user='user', password='VPsystem1234!!', db='vps_planner', charset='utf8', port=1980)   # 데이터 베이스 접속 내용을 conn 변수에 저장
            bom_id = table1.item(row, 0).text()
            order_no = table1.item(row, 2).text()
            part_no = table1.item(row, 5).text()
            dataed = table1.item(row, 15).text()
            partner_company = table1.item(row,18).text()
            division = table1.item(row, 17).text()
            now = datetime.datetime.now()
            take_out_time = now.strftime('%Y-%m-%d %H:%M:%S')
            data = take_out_time.split(' ')[0]
            
            sql = "SELECT qc_time FROM qclist WHERE bom_id = %s AND order_no = %s AND part_no = %s"  # 데이터베이스 명령어를 spl 변수에 저장
            with conn2.cursor() as cur:
                cur.execute(sql,(bom_id,order_no,part_no))
                sqlnowstr = cur.fetchone()

            if sqlnowstr != None:

                sql = "UPDATE qclist SET division = %s, partner_company = %s,take_out_date = %s WHERE bom_id = %s AND order_no = %s AND part_no = %s"   # 데이터베이스 명령어를 spl 변수에 저장
                with conn2.cursor() as cur:
                    cur.execute(sql,(division,partner_company,take_out_time,bom_id,order_no,part_no))
                    conn2.commit()

            elif sqlnowstr == None:
                
                dataed = None
                nowstr = None
                
                sql = "INSERT INTO qclist (bom_id,order_no,part_no,state,qc_time,division,partner_company,take_out_date) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)"   # 데이터베이스 명령어를 spl 변수에 저장
                with conn2.cursor() as cur:
                    cur.execute(sql,(bom_id,order_no,part_no,dataed,nowstr,division,partner_company,take_out_time))
                    conn2.commit()
                print(dataed,2)
        else:
            
            pass
            
    def enter():
        
        if lineEdit_4.text() == '':
            
            lineEdit_4.setFocus()
            
        elif lineEdit_4.text() != '':
            
            qclist.search()
            
    def tableclear():
        
        global table1
        
        table1 = None
        
        table1 = QTableWidget()
        table1.setObjectName("table1")
        table1.setColumnCount(0)
        table1.setRowCount(0)
        gridLayout.addWidget(table1, 2, 0, 1, 1)
        
    def search():
        
        text = lineEdit_4.text()
        
        if text == '':
            
            QMessageBox.information(glowidget,'도번 확인','도번을 입력해 주세요.')
        
        elif text != '':
            
            qclist.plantable(text)
        
    def qcend(complete):
        
        row = table1.currentRow()
        
        if complete == '':
            
            QMessageBox.information(glotab,'수량 입력!','값을 입력해주세요.')
            insertwindow()

        # elif int(complete) > int(table1.item(row,8).text()):
            
        #     QMessageBox.information(glotab,'수량 확인','제작수량보다 많은 수량이 입력되었습니다.\n다시 확인해주세요.')
        #     insertwindow()
            
        elif int(complete) < int(table1.item(row,8).text()):
            
            qclist.qcmenu("수량부족",row,complete)
        
        elif int(complete) >= int(table1.item(row,8).text()):
            
            qclist.qcmenu("검사완료",row,complete)
        
        elif int(complete) == 0:
            
            pass

    def qcmenu(state,row,complete):

        conn2 = pymysql.connect(host='192.168.120.85', user='user', password='VPsystem1234!!', db='vps_planner', charset='utf8', port=1980)   # 데이터 베이스 접속 내용을 conn 변수에 저장
        
        if row == -1:
            row = 0
            
        bom_id = table1.item(row, 0).text()
        order_no = table1.item(row, 2).text()
        part_no = table1.item(row, 5).text()
        quantity = table1.item(row, 8).text()
        now = datetime.datetime.now()
        nowstr = now.strftime('%Y-%m-%d %H:%M:%S')
        division = table1.item(row, 17).text()
        partner_company = table1.item(row, 18).text()
        take_out_time = None
        
        if '검사완료' in state:
            
            dataed = '검사완료 (' + str(complete) + ' / ' + quantity + ')'
            
            
            for i in range(15,17):
                
                if i == 15:
                    
                    table1.item(row, i).setText(dataed)
                    
                elif i == 16:
                    
                    table1.item(row, i).setText(nowstr)

            for j in range(15,17):
                
                table1.item(row, j).setBackground(QColor('#FFCC99'))
                table1.item(row, j).setForeground(QColor('#808080'))
                
            sql = "INSERT INTO qclist (bom_id,order_no,part_no,state,qc_time,division,partner_company,take_out_date) VALUES (%s,%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE state = %s, qc_time = %s, division = %s"   # 데이터베이스 명령어를 spl 변수에 저장
            with conn2.cursor() as cur:
                cur.execute(sql,(bom_id,order_no,part_no,dataed,nowstr,division,partner_company,take_out_time,dataed,nowstr,division))
                conn2.commit()
            print(dataed,3)
        
        elif "수량부족" in state:
            
            dataed = '수량부족 (' + str(complete) + ' / ' + quantity + ')'

            for i in range(15,17):
                
                if i == 15:
                    
                    table1.item(row, i).setText(dataed)
                    
                elif i == 16:
                    
                    table1.item(row, i).setText(nowstr)
            
            for j in range(15,17):     
                
                table1.item(row, j).setBackground(QColor('#FFFF99'))

            sql = "INSERT INTO qclist (bom_id,order_no,part_no,state,qc_time,division,partner_company,take_out_date) VALUES (%s,%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE state = %s, qc_time = %s, division = %s"   # 데이터베이스 명령어를 spl 변수에 저장
            with conn2.cursor() as cur:
                cur.execute(sql,(bom_id,order_no,part_no,dataed,nowstr,division,partner_company,take_out_time,dataed,nowstr,division))
                conn2.commit()
            print(dataed,4)
                    
        elif state == "취소":

            for i in range(15,21):
                
                if table1.item(row, i) != None:
                    
                    table1.item(row, i).setText('')
            
            if row % 2 == 1:
                
                for j in range(20):
                    
                    table1.item(row, j).setForeground(QColor('#000000'))
                    table1.item(row, j).setBackground(QColor('#ebf3ff'))
                    
            elif row % 2 == 0:
                
                for j in range(20):
                    
                    table1.item(row, j).setForeground(QColor('#000000'))
                    table1.item(row, j).setBackground(QColor('#ffffff'))
            
            sql = "DELETE FROM qclist WHERE bom_id = %s AND order_no = %s AND part_no = %s"  # 행삭제 # 데이터베이스 명령어를 spl 변수에 저장
            with conn2.cursor() as cur:
                cur.execute(sql,(bom_id,order_no,part_no))
                conn2.commit()

        elif state == '':

            # sql = "SELECT state,qc_time,division,partner_company,take_out_date FROM qclist WHERE bom_id = %s AND order_no = %s AND part_no = %s"  # 데이터베이스 명령어를 spl 변수에 저장
            # with conn2.cursor() as cur:
            #     cur.execute(sql,(bom_id,order_no,part_no))
            #     state = cur.fetchone()
                
            text = table1.item(0,15).text()

            if text == '':
                
                pass
                

            elif "검사완료" in text:

                for j in range(15,17):

                    table1.item(0, j).setBackground(QColor('#FFCC99'))
                    table1.item(0, j).setForeground(QColor('#808080'))

            elif "검사완료" not in text and text != "":

                for j in range(15,17):

                    table1.item(row, j).setBackground(QColor('#FFFF99'))
    
    def generateMenu(pos):
        
        # 빈공간에서
        if(table1.itemAt(pos) is None):
            
            pass
            
        # 아이템에서
        else:
            
            menu = QMenu()
            row = table1.currentRow()
            column = table1.currentColumn()
            
            if column == 17 or column == 18 or column == 19:
                
                pass
            
            else:
                
                menu.addAction("검사완료", lambda: insertwindow())
                menu.addAction("취소", lambda: qclist.qcmenu("취소",row,0))
                menu.exec_(table1.mapToGlobal(pos))
        
    def plantable(text):   # plantable 함수 시작
        
        start = time.time()
 
        qclist.tableclear()

        table1.setSortingEnabled(True)
        lineEdit_4.clear()

        headerlist = [  'Bom',
                        '  수주일자  ',  
                        ' 관리번호 ',       # 테이블 위젯 헤더에 사용할 리스트 정리 시작    
                        '  고객사  ',
                        ' 어셈블리번호 ',
                        '               파트번호               ',
                        ' 리비전 ',
                        '     품명     ',
                        '  수량  ',
                        '      비고      ',
                        '   제작완료   ',
                        '  후처리완료  ',
                        '   고객납기   ',
                        '              공정              ',
                        '       후처리       ',
                        ' 검사여부\n(완료 / 발주) ',
                        '   검사일시   ',
                        '구분',
                        '  협력사  ',
                        '   반출일시   '
                    ]                     # 테이블 위젯 헤더에 사용할 리스트 정리 끝

        table1.setColumnCount(len(headerlist))   # table1 테이블 위젯 열 갯수를 headerlist 리스트 갯수만큼 설정
        table1.setHorizontalHeaderLabels(headerlist)   # table1 테이블 위젯 열 헤더에 headerlist 리스트 입력
        table1.setAutoScroll(True)   # table1 테이블 위젯 내용이 많아질시 자동 스크롤 생성 허용
        table1.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)   # table1 테이블 위젯 수직 스크롤단위를 픽셀단위로 변경
        table1.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)   # table1 테이블 위젯 수평 스크롤단위를 픽셀단위로 변경
        # table1.setEditTriggers(QAbstractItemView.EditTriggers(False))   # table1 테이블 위젯 셀 수정 불가하게 설정
        table1.setSelectionBehavior(QAbstractItemView.SelectRows)   # table1 테이블 위젯 셀 선택시 행 전체 선택 설정
        table1.setAlternatingRowColors(True)   # table1 테이블 위젯 연속되는 행 색상 다르게 설정
        
        # table1.verticalHeader().hide()
        table1.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        table1.customContextMenuRequested.connect(qclist.generateMenu)
        table1.resizeColumnsToContents()
        table1.resizeRowsToContents()

        conn1 = pymysql.connect(host='152.70.252.118', user='vps_order', password='6006deok!', db='vps_order', charset='utf8', port=3306)   # 데이터 베이스 접속 내용을 conn 변수에 저장
        conn2 = pymysql.connect(host='192.168.120.85', user='user', password='VPsystem1234!!', db='vps_planner', charset='utf8', port=1980)   # 데이터 베이스 접속 내용을 conn 변수에 저장
        start_date = line2.text().replace(' ','') + ' 00:00:01'
        end_date = line3.text().replace(' ','') + ' 23:59:59'
        company_list = []
        
        sql = "SELECT order_date, order_no, customer_name, build_date, retouch_date, delivery_date FROM order_master WHERE create_date BETWEEN %s AND %s"   # 데이터베이스 명령어를 spl 변수에 저장
        with conn1.cursor() as cur:
            cur.execute(sql,(start_date,end_date))
            master = cur.fetchall()
            
        sql = "SELECT bom_id,order_no,part_no,state, qc_time FROM qclist"   # 데이터베이스 명령어를 spl 변수에 저장
        with conn2.cursor() as cur:
            cur.execute(sql)
            state = cur.fetchall()

        for x in master:
            
            order_no = x[1]
             
            sql = "SELECT bom_id, assy_no, parts_no, revision, item_name, quantity, order_detail_desc, routing, retouching FROM order_detail WHERE order_no = %s"   # 데이터베이스 명령어를 spl 변수에 저장
            with conn1.cursor() as cur:
                cur.execute(sql,(order_no))
                detail = cur.fetchall()
                
            for y in detail:
                
                if text in y[2]:
                
                    rowindex = 0
                    data = QTableWidgetItem(str(y[0]))
                    table1.insertRow(rowindex)
                    table1.setItem(rowindex, 0, data)
                    table1.item(rowindex,0).setFlags(table1.item(rowindex,0).flags() ^ Qt.ItemFlag.ItemIsEditable)
                    
                    # table1.item(rowindex,15).setFlags(table1.item(rowindex,15).flags() ^ Qt.ItemFlag.ItemIsEditable)
                    
                    # if state != None:

                    #     if state[1] != None:
                            
                    #         data = QTableWidgetItem(state[1].strftime('%Y/%M/%D'))
                            
                    #     elif state[1] == None:

                    #         data = QTableWidgetItem('')
                        
                    #     table1.setItem(rowindex, 18, data)
                        # 
                    
                    for i in range(1,4):

                        if type(x[i-1]) is datetime.date:
                            
                            data = QTableWidgetItem(x[i-1].strftime('%Y-%m-%d'))
                        
                        elif type(x[i-1]) is str and i != 3:

                            data = QTableWidgetItem(x[i-1])
                        
                        elif i == 3:
                            
                            data = QTableWidgetItem(x[i-1])
                            company_list.append(x[i-1])

                        table1.setItem(rowindex, i, data)
                        table1.item(rowindex,i).setFlags(table1.item(rowindex,i).flags() ^ Qt.ItemFlag.ItemIsEditable)

                    for i in range(4,10):
                        
                        if type(y[i-3]) is int:
                            
                            data = QTableWidgetItem(str(y[i-3]))
                            data.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
                            
                        elif type(y[i-3]) is str:
                            
                            data = QTableWidgetItem(y[i-3])
                        
                        elif y[i-3] == None:

                            data = QTableWidgetItem('')
                        
                        table1.setItem(rowindex, i, data)
                        table1.item(rowindex,i).setFlags(table1.item(rowindex,i).flags() ^ Qt.ItemFlag.ItemIsEditable)
                    
                    for i in range(10,13):
                            
                        data = QTableWidgetItem(x[i-7].strftime('%Y-%m-%d'))

                        table1.setItem(rowindex, i, data)
                        table1.item(rowindex,i).setFlags(table1.item(rowindex,i).flags() ^ Qt.ItemFlag.ItemIsEditable)
                    
                    for i in range(13,15):
                        
                        if y[i-7] != None:
                            
                            data = QTableWidgetItem(str(y[i-6]))
                        
                        elif y[i-7] == None:
                            
                            data = QTableWidgetItem('')
                        
                        table1.setItem(rowindex, i, data)
                        table1.item(rowindex,i).setFlags(table1.item(rowindex,i).flags() ^ Qt.ItemFlag.ItemIsEditable)
                        
                        if i == 13 and '도장' in y[i-6]:
                            
                            table1.item(rowindex, i).setBackground(QColor('#92D050'))
                            
                        elif i == 13 and '도금' in y[i-6]:
                            
                            table1.item(rowindex, i).setBackground(QColor('#00B0F0'))
                    
                    for i in range(15,20):
                        
                        data = QTableWidgetItem('')
                        
                        table1.setItem(rowindex, i, data)
                        
                        if i == 15 or i == 16 or i == 19:

                            table1.item(rowindex,i).setFlags(table1.item(rowindex,i).flags() ^ Qt.ItemFlag.ItemIsEditable)

                    for state_list in state:
                        
                        if x[1] in state_list and y[0] in state_list and y[2] in state_list and state_list[3] is not None and state_list[4] is not None:
                            
                            stateitem = QTableWidgetItem(state_list[3])
                            statedate = QTableWidgetItem(state_list[4].strftime('%Y-%m-%d'))
                            table1.setItem(rowindex, 15, stateitem) 
                            table1.item(rowindex,15).setFlags(table1.item(rowindex,15).flags() ^ Qt.ItemFlag.ItemIsEditable)
                            table1.setItem(rowindex, 16, statedate) 
                            table1.item(rowindex,16).setFlags(table1.item(rowindex,16).flags() ^ Qt.ItemFlag.ItemIsEditable)

                            if state_list[3] == '':
                        
                                pass
                                

                            elif "검사완료" in state_list[3]:

                                for j in range(20):

                                    table1.item(rowindex, j).setBackground(QColor('#FFCC99'))
                                    table1.item(rowindex, j).setForeground(QColor('#808080'))

                            elif "검사완료" not in state_list[3] and state_list[3] != "":

                                for j in range(20):

                                    table1.item(rowindex, j).setBackground(QColor('#FFFF99'))
                            
                else:
                    
                    pass
        
        hidec = [0,4,9]
            
        for xx in hidec:
            
            table1.hideColumn(xx)
        
        table1.itemChanged.connect(qclist.changeitem)
        
        print("time :", time.time() - start)

class insertwindow(QMainWindow,form_insertclass):

    def __init__(self):

        super().__init__()
        self.setWindowModality(Qt.ApplicationModal)
        self.ui = uic.loadUi("resource/ui/insert.ui",self)
        self.show()
        self.b_lineEdit.setFocus()
        self.b_lineEdit.setValidator(QIntValidator())
        self.b_pushButton.clicked.connect(self.sendcommand)
        self.b_pushButton_2.clicked.connect(self.closes)

    def sendcommand(self):
        
        msg1 = self.b_lineEdit.text()
        qclist.qcend(msg1)
        self.close()

    def closes(self):

        self.close()

    def keyPressEvent(self,e):

        if e.key() == Qt.Key_Escape:
            self.close()
        if e.key() == Qt.Key_Return:
            self.sendcommand()
        if e.key() == Qt.Key_Enter:
            self.sendcommand()